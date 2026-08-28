from datetime import date, datetime, timedelta, time as dt_time
from functools import wraps
from io import BytesIO
import hashlib, ipaddress, json, os, secrets, io, re, hmac, base64, gzip, time as time_module
from pathlib import Path
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
from flask import Flask, abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for, g
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash
from sqlalchemy import func, or_, and_, text, cast, Integer, case
from sqlalchemy.exc import IntegrityError

try:
    import numpy as np
    import cv2
except Exception:
    np = None
    cv2 = None

from models import (db, Account, Student, Parent, ParentStudent, Teacher, TeacherAssignment,
    TeacherSubjectAssignment, Subject, Exam, Mark, AssessmentComponent, Attendance,
    SchoolCalendar, Announcement, ResultPublication, AuditEvent, SchoolSetting, SchoolClock,
    PublishedReport, ReportCardConfig, FeeStructure, FeePaymentWindow, FeeStructureDocument,
    FeeInvoice, FeePayment)
from exports import build_xlsx, build_pdf, build_report_card
from face_utils import (available as face_available, encode_frame, best_match, match_distance,
    decode_data_url, recognize_faces, image_quality, best_match_for_encoding)

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

try:
    import pytesseract
except Exception:
    pytesseract = None

load_dotenv()
BASE_DIR = Path(__file__).resolve().parent
app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.getenv('SECRET_KEY', secrets.token_hex(32)),
    SQLALCHEMY_DATABASE_URI=os.getenv('DATABASE_URL','sqlite:///'+str(BASE_DIR/'school.db')).replace('postgres://','postgresql://'),
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    SQLALCHEMY_ENGINE_OPTIONS={
        'pool_pre_ping': True,
        'pool_recycle': 280,
        'pool_size': int(os.getenv('DB_POOL_SIZE','2')),
        'max_overflow': int(os.getenv('DB_MAX_OVERFLOW','2')),
        'pool_timeout': int(os.getenv('DB_POOL_TIMEOUT','10')),
    },
    JWT_SECRET_KEY=os.getenv('JWT_SECRET_KEY') or os.getenv('SECRET_KEY') or secrets.token_hex(32),
    JWT_ACCESS_TTL_MINUTES=int(os.getenv('JWT_ACCESS_TTL_MINUTES','30')),
    JWT_REFRESH_TTL_DAYS=int(os.getenv('JWT_REFRESH_TTL_DAYS','7')),
    CSRF_TTL_MINUTES=int(os.getenv('CSRF_TTL_MINUTES','30')),
    CSRF_GRACE_MINUTES=int(os.getenv('CSRF_GRACE_MINUTES','10')),
    APP_TIMEZONE=os.getenv('APP_TIMEZONE','Asia/Kolkata'),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=os.getenv('SESSION_COOKIE_SECURE', 'true' if os.getenv('LOCAL_HTTPS','false').lower() in {'1','true','yes','on'} else 'false').lower()=='true',
    LOCAL_HTTPS=os.getenv('LOCAL_HTTPS','false').lower() in {'1','true','yes','on'},
    SSL_CERT_FILE=os.getenv('SSL_CERT_FILE',''),
    SSL_KEY_FILE=os.getenv('SSL_KEY_FILE',''),
    SSL_ADHOC=os.getenv('SSL_ADHOC','false').lower() in {'1','true','yes','on'},
)
# Small local SQLite tuning: fewer lock errors and better concurrent reads.
if app.config['SQLALCHEMY_DATABASE_URI'].startswith('sqlite:///'):
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        **app.config.get('SQLALCHEMY_ENGINE_OPTIONS', {}),
        'connect_args': {'check_same_thread': False, 'timeout': 30},
    }

if os.getenv('TRUST_PROXY_HEADERS','false').lower()=='true':
    # Render terminates TLS at its load balancer and forwards the original client IP.
    # Only enable this when the app is actually behind a trusted single proxy hop.
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
db.init_app(app)


def _request_cache():
    """Return a tiny per-request cache without using global mutable state."""
    cache = getattr(g, '_school_request_cache', None)
    if cache is None:
        cache = {}
        setattr(g, '_school_request_cache', cache)
    return cache


def csv_values(name):
    return [x.strip() for x in os.getenv(name,'').split(',') if x.strip()]

def client_ip():
    return request.remote_addr or 'unknown'

def ip_allowed(name):
    allowed = csv_values(name)
    if not allowed: return True
    ip = ipaddress.ip_address(client_ip())
    for item in allowed:
        try:
            if ip == ipaddress.ip_address(item): return True
        except ValueError:
            try:
                if ip in ipaddress.ip_network(item, strict=False): return True
            except ValueError: pass
    return False


def _jwt_decode(token, expected_type='access'):
    if not token:
        return None
    try:
        import jwt
        payload=jwt.decode(token, app.config['JWT_SECRET_KEY'], algorithms=['HS256'])
        if payload.get('type') != expected_type:
            return None
        return payload
    except Exception:
        return None


def _jwt_payload_from_request():
    auth=request.headers.get('Authorization','')
    if not auth.lower().startswith('bearer '):
        return None
    token=auth.split(None,1)[1].strip()
    return _jwt_decode(token, 'access')


def jwt_account_from_request():
    payload=_jwt_payload_from_request()
    if not payload:
        return None
    try:
        aid=int(payload.get('sub'))
    except (TypeError, ValueError):
        return None
    acct=db.session.get(Account, aid)
    if not acct or not acct.active:
        return None
    g.jwt_authenticated=True
    return acct


def _issue_jwt(acct):
    import jwt
    now=datetime.utcnow()
    exp=now+timedelta(minutes=app.config['JWT_ACCESS_TTL_MINUTES'])
    return jwt.encode({'sub':str(acct.id),'role':acct.role,'type':'access','iat':int(now.timestamp()),'exp':int(exp.timestamp()),'jti':secrets.token_hex(12)}, app.config['JWT_SECRET_KEY'], algorithm='HS256')


def _issue_refresh_jwt(acct):
    import jwt
    now=datetime.utcnow()
    exp=now+timedelta(days=app.config['JWT_REFRESH_TTL_DAYS'])
    return jwt.encode({'sub':str(acct.id),'role':acct.role,'type':'refresh','iat':int(now.timestamp()),'exp':int(exp.timestamp()),'jti':secrets.token_hex(16)}, app.config['JWT_SECRET_KEY'], algorithm='HS256')


def _ensure_csrf_token(force=False):
    now_ts=time_module.time()
    current=session.get('csrf')
    issued=float(session.get('csrf_issued_at',0) or 0)
    ttl=app.config['CSRF_TTL_MINUTES']*60
    if force or not current or now_ts-issued >= ttl:
        if current:
            session['csrf_previous']=current
            session['csrf_previous_issued_at']=issued
        session['csrf']=secrets.token_urlsafe(32)
        session['csrf_issued_at']=now_ts
    return session['csrf']


def _csrf_is_valid(supplied):
    if not supplied:
        return False
    supplied=str(supplied)
    current=str(session.get('csrf') or '')
    if current and hmac.compare_digest(supplied,current):
        return True
    previous=str(session.get('csrf_previous') or '')
    prev_issued=float(session.get('csrf_previous_issued_at',0) or 0)
    grace=app.config['CSRF_GRACE_MINUTES']*60
    return bool(previous and hmac.compare_digest(supplied,previous) and time_module.time()-prev_issued <= grace)


@app.before_request
def _csrf_guard():
    # Every browser session has a renewable CSRF token. A short grace period
    # keeps already-open forms usable while the token refreshes in the background.
    _ensure_csrf_token()
    if request.method not in {'POST','PUT','PATCH','DELETE'}:
        return None
    if request.path in {'/login','/api/auth/login','/api/auth/refresh','/healthz'}:
        return None
    # Bearer-token API calls are authenticated independently of browser CSRF.
    if request.headers.get('Authorization','').lower().startswith('bearer '):
        return None
    supplied=request.headers.get('X-CSRF-Token') or request.form.get('_csrf')
    if not _csrf_is_valid(supplied):
        if request.path.startswith('/api/'):
            return jsonify({'error':'Invalid or expired CSRF token. Refresh the page/token and try again.','csrf_refresh':url_for('csrf_token_api')}),400
        return ('Invalid or expired CSRF token. Refresh the page and try again.',400)
    return None


@app.get('/api/csrf')
def csrf_token_api():
    token=_ensure_csrf_token()
    issued=float(session.get('csrf_issued_at',time_module.time()) or time_module.time())
    remaining=max(0,int(app.config['CSRF_TTL_MINUTES']*60-(time_module.time()-issued)))
    return jsonify({'csrf_token':token,'expires_in':remaining})


@app.post('/api/auth/login')
def api_auth_login():
    data=request.get_json(silent=True) or request.form
    username=str(data.get('username','')).strip()
    password=str(data.get('password',''))
    acct=Account.query.filter(func.lower(Account.username)==username.lower()).first()
    if not acct or not acct.active or not check_password_hash(acct.password_hash,password):
        return jsonify({'error':'Invalid username or password.'}),401
    acct.last_login=datetime.utcnow()
    db.session.commit()
    return jsonify({'access_token':_issue_jwt(acct),'refresh_token':_issue_refresh_jwt(acct),'token_type':'Bearer','expires_in':app.config['JWT_ACCESS_TTL_MINUTES']*60,'refresh_expires_in':app.config['JWT_REFRESH_TTL_DAYS']*86400,'account':{'id':acct.id,'username':acct.username,'role':acct.role,'display_name':acct.display_name}})


@app.post('/api/auth/refresh')
def api_auth_refresh():
    data=request.get_json(silent=True) or request.form
    token=str(data.get('refresh_token') or '').strip()
    payload=_jwt_decode(token,'refresh')
    if not payload:
        return jsonify({'error':'Invalid or expired refresh token.'}),401
    try:
        aid=int(payload.get('sub'))
    except (TypeError,ValueError):
        return jsonify({'error':'Invalid refresh token subject.'}),401
    acct=db.session.get(Account,aid)
    if not acct or not acct.active:
        return jsonify({'error':'Account is inactive or unavailable.'}),401
    return jsonify({'access_token':_issue_jwt(acct),'refresh_token':_issue_refresh_jwt(acct),'token_type':'Bearer','expires_in':app.config['JWT_ACCESS_TTL_MINUTES']*60,'refresh_expires_in':app.config['JWT_REFRESH_TTL_DAYS']*86400})


def staff_network_required():
    def deco(fn):
        @wraps(fn)
        def wrapper(*args,**kwargs):
            if not ip_allowed('STAFF_ALLOWED_IPS'):
                return ('<h1>403</h1><p>Staff services are available only from an authorized school network.</p>',403)
            return fn(*args,**kwargs)
        return wrapper
    return deco

def audit_network_required():
    def deco(fn):
        @wraps(fn)
        def wrapper(*args,**kwargs):
            if not ip_allowed('AUDIT_LOG_ALLOWED_IPS'):
                return ('<h1>403</h1><p>Restricted audit service.</p>',403)
            return fn(*args,**kwargs)
        return wrapper
    return deco

def login_required(role=None):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args,**kwargs):
            acct=None
            if request.headers.get('Authorization','').lower().startswith('bearer '):
                acct=jwt_account_from_request()
            if not acct:
                aid=session.get('account_id')
                if aid:
                    acct=db.session.get(Account, aid)
            if not acct or not acct.active:
                if request.path.startswith('/api/'):
                    return jsonify({'error':'Authentication required.'}),401
                return redirect(url_for('login', next=request.path))
            g.current_account=acct
            if role and acct.role != role: abort(403)
            if acct.must_change_password and request.endpoint not in {'change_credentials','logout'} and not g.get('jwt_authenticated'):
                return redirect(url_for('change_credentials'))
            return fn(*args,**kwargs)
        return wrapper
    return deco

def staff_required(fn):
    @wraps(fn)
    @staff_network_required()
    @login_required()
    def wrapper(*args,**kwargs):
        acct=current_account();
        if not acct or acct.role not in {'admin','teacher'}: abort(403)
        return fn(*args,**kwargs)
    return wrapper

def admin_required(fn):
    @wraps(fn)
    @staff_network_required()
    @login_required('admin')
    def wrapper(*args,**kwargs): return fn(*args,**kwargs)
    return wrapper

@app.after_request
def _performance_headers(response):
    if request.path.startswith('/static/'):
        response.headers.setdefault('Cache-Control', 'public, max-age=86400, stale-while-revalidate=604800')
    try:
        response.headers['X-CSRF-Token']=session.get('csrf', '')
    except RuntimeError:
        pass
    return response

@app.after_request
def _gzip_response(response):
    # Compress larger HTML/JSON/CSS/JS responses to make remote-phone access snappier.
    if response.status_code in (204, 304) or response.headers.get('Content-Encoding') or response.direct_passthrough:
        return response
    if 'gzip' not in (request.headers.get('Accept-Encoding') or '').lower():
        return response
    content_type=response.headers.get('Content-Type','')
    if not any(content_type.startswith(t) for t in ('text/','application/json','application/javascript')):
        return response
    try:
        body=response.get_data()
    except Exception:
        return response
    if len(body) < 1024:
        return response
    compressed=gzip.compress(body, compresslevel=5)
    if len(compressed) >= len(body):
        return response
    response.set_data(compressed)
    response.headers['Content-Encoding']='gzip'
    response.headers['Vary']='Accept-Encoding'
    response.headers['Content-Length']=str(len(compressed))
    return response

def current_account():
    acct=getattr(g,'current_account',None)
    if acct: return acct
    if hasattr(g, '_current_account_loaded'):
        return getattr(g, 'current_account', None)
    aid=session.get('account_id')
    acct=db.session.get(Account, aid) if aid else None
    g.current_account=acct
    g._current_account_loaded=True
    return acct


def log_audit(action, target_type='', target_id='', extra=None):
    acct=current_account(); actor=acct.username if acct else session.get('audit_username','system'); role=acct.role if acct else 'audit'
    prev=db.session.query(AuditEvent).order_by(AuditEvent.id.desc()).first()
    prev_hash=prev.event_hash if prev else ''
    payload={'action':action,'target_type':target_type,'target_id':str(target_id or ''),'ip':client_ip(),'at':datetime.utcnow().isoformat(),'extra':extra or {}}
    raw=prev_hash+json.dumps(payload,sort_keys=True,separators=(',',':'))
    h=hashlib.sha256(raw.encode()).hexdigest()
    db.session.add(AuditEvent(actor_username=actor,actor_role=role,action=action,target_type=target_type,target_id=str(target_id or ''),ip_address=client_ip(),user_agent=request.headers.get('User-Agent',''),metadata_json=json.dumps(extra or {}),previous_hash=prev_hash,event_hash=h))


def normalize_school_name(value, field='Name'):
    raw=' '.join(str(value or '').strip().split()).upper()
    return re.sub(r"[^A-Z0-9 .'\-]", '', raw)[:160]

ROMAN_TO_INT = {'I':1,'II':2,'III':3,'IV':4,'V':5,'VI':6,'VII':7,'VIII':8,'IX':9,'X':10,'XI':11,'XII':12}
INT_TO_ROMAN = {v:k for k,v in ROMAN_TO_INT.items()}

def class_number(value):
    raw=''.join(str(value or '').strip().upper().split())
    if raw.isdigit():
        n=int(raw)
    else:
        n=ROMAN_TO_INT.get(raw, 0)
    return n

def normalize_class(value):
    n=class_number(value)
    if 1 <= n <= 10:
        return INT_TO_ROMAN[n]
    raw=' '.join(str(value or '').strip().split()).upper()
    return re.sub(r'[^A-Z0-9 \-]', '', raw)[:10]

def normalize_section(value):
    raw=' '.join(str(value or '').strip().split()).upper()
    return re.sub(r'[^A-Z0-9]', '', raw)[:10]

def parse_bool(value, default=False):
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    raw=str(value).strip().lower()
    if raw in {'1','true','yes','on','working','present'}:
        return True
    if raw in {'0','false','no','off','holiday','non-working','nonworking'}:
        return False
    return default


def release_inactive_account_username(username):
    """Free a previously deleted/inactive username while preserving the person record.

    Older builds soft-deactivated accounts, leaving the UNIQUE(username) constraint
    occupied forever. Detaching the account from its person before deletion preserves
    the student/teacher/parent record while making the username reusable.
    """
    normalized=(username or '').strip().lower()
    if not normalized:
        return False
    acct=Account.query.filter(func.lower(Account.username)==normalized).first()
    if not acct or acct.active:
        return False
    Student.query.filter_by(account_id=acct.id).update({Student.account_id: None}, synchronize_session=False)
    Teacher.query.filter_by(account_id=acct.id).update({Teacher.account_id: None}, synchronize_session=False)
    Parent.query.filter_by(account_id=acct.id).update({Parent.account_id: None}, synchronize_session=False)
    db.session.delete(acct)
    db.session.flush()
    return True

def student_order(query):
    """Stable school ordering: class number -> section -> natural roll -> name.

    Nonnumeric/legacy roll numbers are handled safely without a PostgreSQL cast
    that can abort the entire query. Numeric rolls still sort naturally (1,2,10).
    """
    class_rank = case(
        {'NURSERY': 0, 'LKG': 1, 'UKG': 2, **{k:v+2 for k,v in ROMAN_TO_INT.items()}},
        value=func.upper(Student.class_name),
        else_=99,
    )
    roll_text=func.trim(Student.roll_number)
    numeric_only = roll_text.op('~')('^[0-9]+$')
    roll_num = case((numeric_only, cast(roll_text, Integer)), else_=2147483647)
    return query.order_by(class_rank, Student.class_name, Student.section, roll_num, roll_text, Student.name, Student.id)

def language_label(code):
    return {'telugu':'Telugu','hindi':'Hindi','sanskrit':'Sanskrit'}.get(str(code or '').lower(), str(code or '').title())

def subject_options_for_class(class_name, second_language=None, third_language=None):
    n=class_number(class_name)
    if 5<=n<=8:
        return [('eng','English (1st Language)'),
                (f'lang2_{str(second_language).lower()}','2nd Language: '+language_label(second_language)) if second_language else ('lang2_placeholder','2nd Language'),
                (f'lang3_{str(third_language).lower()}','3rd Language: '+language_label(third_language)) if third_language else ('lang3_placeholder','3rd Language'),
                ('computers','Computers'),('math','Mathematics'),('social','Social Science'),('science','Science')]
    if 9<=n<=10:
        return [('math','Mathematics'),('chemistry','Chemistry'),('biology','Biology'),('physics','Physics'),('social','Social Science'),('eng','English (1st Language)'),
                (f'lang2_{str(second_language).lower()}','2nd Language: '+language_label(second_language)) if second_language else ('lang2_placeholder','2nd Language'),('it','Information Technology')]
    return []

def class_list():
    return ['NURSERY','LKG','UKG','I','II','III','IV','V','VI','VII','VIII','IX','X','XI','XII']

def seed_data():
    if not db.session.query(Account).filter_by(role='admin').first():
        u=os.getenv('INITIAL_ADMIN_USERNAME','admin'); p=os.getenv('INITIAL_ADMIN_PASSWORD','ChangeThisImmediately123!')
        db.session.add(Account(username=u,password_hash=generate_password_hash(p),role='admin',display_name='SCHOOL ADMINISTRATOR',must_change_password=True))
    if not Subject.query.first():
        subjects=[]
        base_common=[('eng','English'),('math','Mathematics')]
        for code,name in base_common: subjects.append(Subject(code=code,name=name,class_band='5-10'))
        for code,name in [('social','Social Science'),('science','Science'),('computers','Computers'),('it','Information Technology'),('physics','Physics'),('chemistry','Chemistry'),('biology','Biology')]: subjects.append(Subject(code=code,name=name,class_band='5-10'))
        for lang in ['telugu','hindi','sanskrit']:
            subjects.append(Subject(code=f'lang2_{lang}',name=lang.title(),class_band='5-10',language_group='second'))
            subjects.append(Subject(code=f'lang3_{lang}',name=lang.title(),class_band='5-8',language_group='third'))
        db.session.add_all(subjects)
    # Normalize language subject names so identical languages remain visibly distinct
    # between 2nd and 3rd language slots.
    for code, label in [
        ('lang2_telugu','Telugu (2nd Language)'),
        ('lang2_hindi','Hindi (2nd Language)'),
        ('lang2_sanskrit','Sanskrit (2nd Language)'),
        ('lang3_telugu','Telugu (3rd Language)'),
        ('lang3_hindi','Hindi (3rd Language)'),
        ('lang3_sanskrit','Sanskrit (3rd Language)')
    ]:
        subj=Subject.query.filter_by(code=code).first()
        if subj: subj.name=label
    if not Exam.query.first():
        db.session.add_all([Exam(name='PT-1',max_marks=40,order_index=1),Exam(name='PT-2',max_marks=80,order_index=2),Exam(name='PT-3',max_marks=40,order_index=3),Exam(name='Final Examination',max_marks=80,order_index=4,is_final=True)])
    if SchoolSetting.query.filter_by(key='academic_session').first() is None:
        db.session.add(SchoolSetting(key='academic_session',value=os.getenv('ACADEMIC_SESSION','2026-27')))
    # Backfill missing accounts for legacy students so every student record has a login.
    legacy_password=os.getenv('LEGACY_STUDENT_DEFAULT_PASSWORD','ChangeThisImmediately123!')
    for st in Student.query.filter(Student.account_id.is_(None), Student.active.is_(True)).limit(5000).all():
        base=''.join(ch.lower() if ch.isalnum() else '_' for ch in (st.admission_number or f'student_{st.id}')).strip('_') or f'student_{st.id}'
        username=base
        n=1
        while Account.query.filter_by(username=username).first():
            n+=1; username=f'{base}_{n}'
        acct=Account(username=username,password_hash=generate_password_hash(legacy_password),role='student',display_name=st.name,must_change_password=True,active=st.active)
        db.session.add(acct); db.session.flush(); st.account_id=acct.id
    for t in Teacher.query.filter(Teacher.account_id.is_(None), Teacher.active.is_(True)).limit(1000).all():
        base=''.join(ch.lower() if ch.isalnum() else '_' for ch in (t.name or f'teacher_{t.id}')).strip('_') or f'teacher_{t.id}'
        username=base; n=1
        while Account.query.filter_by(username=username).first():
            n+=1; username=f'{base}_{n}'
        acct=Account(username=username,password_hash=generate_password_hash(legacy_password),role='teacher',display_name=t.name,must_change_password=True,active=t.active)
        db.session.add(acct); db.session.flush(); t.account_id=acct.id
    db.session.commit()



def ensure_announcement_schema():
    """Ensure targeted announcement column/index exists on older databases."""
    try:
        db.session.execute(text("ALTER TABLE announcement ADD COLUMN IF NOT EXISTS parent_id INTEGER REFERENCES parent(id) ON DELETE CASCADE"))
        db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_announcement_parent ON announcement(parent_id, published)"))
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise


def grade_for_percent(pct):
    try: pct=float(pct)
    except Exception: return 'E'
    if pct >= 91: return 'A1'
    if pct >= 81: return 'A2'
    if pct >= 71: return 'B1'
    if pct >= 61: return 'B2'
    if pct >= 51: return 'C1'
    if pct >= 41: return 'C2'
    if pct >= 33: return 'D'
    return 'E'

def pass_fail_for_percent(pct):
    return 'PASS' if float(pct) >= 33 else 'REQUIRES ATTENTION'

def subjects_for_class(class_name, second_language=None, third_language=None):
    codes=[c for c,_ in subject_options_for_class(class_name,second_language,third_language) if not c.endswith('_placeholder')]
    if not codes:
        return []
    by_code={s.code:s for s in Subject.query.filter(Subject.code.in_(codes)).all()}
    return [by_code[c] for c in codes if c in by_code]

def working_days_between(start,end):
    rows=SchoolCalendar.query.filter(SchoolCalendar.date>=start,SchoolCalendar.date<=end).all()
    overrides={r.date:r.is_working for r in rows}
    cur=start; count=0
    while cur<=end:
        count += 1 if overrides.get(cur,cur.weekday()<5) else 0
        cur += timedelta(days=1)
    return count

def attendance_summary(student_id, start, end):
    wd=working_days_between(start,end)
    ats=Attendance.query.filter_by(student_id=student_id).filter(Attendance.date>=start,Attendance.date<=end).all()
    present=sum(1 for a in ats if a.status in {'present','late'})
    absent=sum(1 for a in ats if a.status=='absent')
    pct=round((present/wd)*100,2) if wd else 0
    return {'working_days':wd,'present':present,'absent':absent,'percentage':pct}

def attendance_percentage_map(student_ids, start, end):
    ids=[int(x) for x in student_ids if x is not None]
    if not ids:
        return {}
    wd=working_days_between(start,end)
    if wd <= 0:
        return {sid:0 for sid in ids}
    rows=(db.session.query(Attendance.student_id, func.count(Attendance.id))
          .filter(Attendance.student_id.in_(ids), Attendance.date>=start, Attendance.date<=end, Attendance.status.in_(['present','late']))
          .group_by(Attendance.student_id).all())
    return {sid: round((count/wd)*100,2) for sid,count in rows}

def school_timezone():
    zone_name=app.config.get('APP_TIMEZONE','Asia/Kolkata')
    try:
        return ZoneInfo(zone_name)
    except Exception:
        # Windows installations can lack the IANA tzdata package. Keep the
        # school clock functional even then; requirements.txt also includes tzdata.
        if zone_name == 'Asia/Kolkata':
            return __import__('datetime').timezone(timedelta(hours=5, minutes=30), 'IST')
        return __import__('datetime').timezone.utc

def _configured_clock_time():
    row=db.session.get(SchoolClock,1)
    raw=(row.override_time or '').strip() if row else ''
    if not raw:
        return None
    for fmt in ('%H:%M','%H:%M:%S'):
        try:
            return datetime.strptime(raw,fmt).time()
        except ValueError:
            pass
    return None

def school_now():
    now=datetime.now(school_timezone())
    override=_configured_clock_time()
    if override is not None:
        return datetime.combine(now.date(),override,tzinfo=school_timezone())
    return now

def school_time(): return school_now().time().replace(microsecond=0)

def school_date(): return school_now().date()

def is_working_day(day):
    override = SchoolCalendar.query.filter_by(date=day).first()
    if override is not None:
        return bool(override.is_working)
    return day.weekday() < 5


def weekly_default_is_working(day):
    return day.weekday() < 5


def get_school_clock_override():
    row = db.session.get(SchoolClock, 1)
    return (row.override_time or '').strip() if row else None


def json_error(message, status=400):
    return jsonify({'error': str(message)}), status


def attendance_status_for_time(value=None):
    t=value or school_time()
    def parse_env(name, default):
        raw=os.getenv(name,default)
        try: return datetime.strptime(raw,'%H:%M').time()
        except ValueError: return datetime.strptime(default,'%H:%M').time()
    present_from=parse_env('ATTENDANCE_PRESENT_FROM','07:30')
    late_after=parse_env('ATTENDANCE_LATE_AFTER','08:30')
    absent_after=parse_env('ATTENDANCE_ABSENT_AFTER','09:00')
    if t < present_from: return 'not_open'
    if t < late_after: return 'present'
    if t < absent_after: return 'late'
    return 'closed'

def school_year_bounds(): return date(school_date().year if school_date().month>=4 else school_date().year-1,4,1), date(school_date().year+1 if school_date().month>=4 else school_date().year,3,31)

def allowed_students_for_account(acct):
    if acct.role=='admin': return Student.query.filter_by(active=True)
    if acct.role=='teacher':
        t=Teacher.query.filter_by(account_id=acct.id).first()
        if not t: return Student.query.filter(False)
        assignments=TeacherAssignment.query.filter_by(teacher_id=t.id).all()
        from sqlalchemy import or_, and_
        clauses=[and_(Student.class_name==a.class_name,Student.section==a.section) for a in assignments]
        return Student.query.filter(or_(*clauses)) if clauses else Student.query.filter(False)
    if acct.role=='student':
        s=Student.query.filter_by(account_id=acct.id).first(); return Student.query.filter_by(id=s.id) if s else Student.query.filter(False)
    if acct.role=='parent':
        p=Parent.query.filter_by(account_id=acct.id).first(); ids=[x.student_id for x in ParentStudent.query.filter_by(parent_id=p.id).all()] if p else []
        return Student.query.filter(Student.id.in_(ids)) if ids else Student.query.filter(False)
    return Student.query.filter(False)

@app.get('/healthz')
def healthz():
    try:
        db.session.execute(text('SELECT 1'))
        return jsonify({'status':'ok','database':'ok'})
    except Exception:
        db.session.rollback()
        return jsonify({'status':'degraded','database':'error'}),503


def _local_ai_fallback(question, acct, students, attendance_pct, academic_pct):
    q=question.lower().strip()
    # Deterministic, scope-safe school assistant. It never depends on an API quota.
    if 'how many students' in q and ('attendance' in q or 'absent' in q):
        m=re.search(r'(?:less than|below|under)\s+(\d+(?:\.\d+)?)',q)
        if m:
            t=float(m.group(1)); n=sum(1 for s in students if attendance_pct.get(s.id,0)<t)
            return f'{n} student{"s" if n!=1 else ""} in your allowed scope have attendance below {t:g}%. '
    if ('which students' in q or 'who' in q or 'list' in q) and ('attendance' in q or 'absent' in q):
        m=re.search(r'(?:less than|below|under)\s+(\d+(?:\.\d+)?)',q)
        if m:
            t=float(m.group(1)); hits=[s for s in students if attendance_pct.get(s.id,0)<t]
            return f'\n'.join([f'{len(hits)} students match your attendance filter.']+[f'{s.name} ({s.class_name}{("-"+s.section) if s.section else ""}) — {attendance_pct.get(s.id,0):g}%' for s in hits[:100]]) if hits else 'No students match your attendance filter.'
    if ('which students' in q or 'who' in q or 'list' in q) and ('academic' in q or 'academics' in q or 'marks' in q or 'results' in q):
        m=re.search(r'(?:less than|below|under)\s+(\d+(?:\.\d+)?)',q)
        if m:
            t=float(m.group(1)); hits=[s for s in students if academic_pct.get(s.id,0)<t]
            return '\n'.join([f'{len(hits)} students match your academic filter.']+[f'{s.name} ({s.class_name}{("-"+s.section) if s.section else ""}) — {academic_pct.get(s.id,0):g}%' for s in hits[:100]]) if hits else 'No students match your academic filter.'
    if 'how many students' in q and ('academic' in q or 'academics' in q or 'marks' in q or 'results' in q):
        m=re.search(r'(?:less than|below|under)\s+(\d+(?:\.\d+)?)',q)
        if m:
            t=float(m.group(1)); n=sum(1 for s in students if academic_pct.get(s.id,0)<t)
            return f'{n} student{"s" if n!=1 else ""} in your allowed scope have academic performance below {t:g}%. '
    if 'how many students' in q:
        return f'You can access {len(students)} student records in your current scope.'
    return 'I can answer questions using the school data available to your account. Try asking about attendance, academics/results, students, classes, or parents.'


@app.get('/announcements')
def announcements_view():
    acct = current_account()
    if not acct:
        q = Announcement.query.filter(Announcement.published.is_(True), Announcement.audience.in_(['public','all']))
    else:
        role_audience = {'teacher':'teachers','student':'students','parent':'parents'}.get(acct.role)
        if acct.role == 'admin':
            q = Announcement.query.filter(Announcement.published.is_(True), Announcement.audience.in_(['public','all','admin','teachers','students','parents']))
        elif acct.role == 'parent':
            parent = Parent.query.filter_by(account_id=acct.id).first()
            clauses=[Announcement.audience.in_(['public','all']), and_(Announcement.audience=='parents', Announcement.parent_id.is_(None))]
            if parent:
                clauses.append(Announcement.parent_id==parent.id)
            q=Announcement.query.filter(Announcement.published.is_(True), or_(*clauses))
        else:
            clauses=[Announcement.audience.in_(['public','all'])]
            if role_audience:
                clauses.append(Announcement.audience==role_audience)
            q=Announcement.query.filter(Announcement.published.is_(True), or_(*clauses))
    announcements = q.order_by(Announcement.published_at.desc(), Announcement.id.desc()).limit(100).all()
    return render_template('announcements_public.html', announcements=announcements)

@app.get('/')
def index():
    if session.get('account_id'): return redirect(url_for('dashboard'))
    return render_template('landing.html')

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method=='POST':
        username=request.form.get('username','').strip(); password=request.form.get('password','')
        acct=Account.query.filter(func.lower(Account.username)==username.lower()).first()
        if not acct or not acct.active or not check_password_hash(acct.password_hash,password):
            return render_template('login.html',error='Invalid username or password.'),401
        session.clear(); session['account_id']=acct.id; session['role']=acct.role; session['csrf']=secrets.token_urlsafe(32); session['csrf_issued_at']=time_module.time(); session.pop('csrf_previous',None); session.pop('csrf_previous_issued_at',None); acct.last_login=datetime.utcnow(); log_audit('login', 'Account', acct.id); db.session.commit()
        next_url=request.args.get('next','') or ''
        if not next_url.startswith('/') or next_url.startswith('//'):
            next_url=url_for('dashboard')
        return redirect(next_url)
    return render_template('login.html')

@app.get('/logout')
def logout(): session.clear(); return redirect(url_for('index'))

PROFILE_DIR = BASE_DIR / 'static' / 'uploads' / 'profiles'
PROFILE_DIR.mkdir(parents=True, exist_ok=True)

def profile_picture_url(account):
    if not account or not account.profile_picture_path:
        return None
    return url_for('static', filename=account.profile_picture_path)

@app.post('/account/profile-picture')
@login_required()
def upload_profile_picture():
    acct=current_account()
    upload=request.files.get('profile_picture')
    if not upload or not upload.filename:
        flash('CHOOSE A PROFILE PICTURE FIRST.','error')
        return redirect(url_for('change_credentials'))
    raw=upload.read()
    if len(raw) > 3 * 1024 * 1024:
        flash('PROFILE PICTURE MUST BE 3 MB OR SMALLER.','error')
        return redirect(url_for('change_credentials'))
    if np is None or cv2 is None:
        flash('PROFILE PICTURES REQUIRE NUMPY AND OPENCV.','error')
        return redirect(url_for('change_credentials'))
    arr=np.frombuffer(raw,dtype=np.uint8)
    img=cv2.imdecode(arr,cv2.IMREAD_COLOR)
    if img is None:
        flash('INVALID IMAGE. USE JPG, PNG OR WEBP.','error')
        return redirect(url_for('change_credentials'))
    h,w=img.shape[:2]
    if h < 64 or w < 64:
        flash('PROFILE PICTURE IS TOO SMALL.','error')
        return redirect(url_for('change_credentials'))
    side=min(h,w)
    y=(h-side)//2; x=(w-side)//2
    img=img[y:y+side,x:x+side]
    img=cv2.resize(img,(512,512),interpolation=cv2.INTER_AREA)
    ok,encoded=cv2.imencode('.jpg',img,[int(cv2.IMWRITE_JPEG_QUALITY),90])
    if not ok:
        flash('COULD NOT PROCESS THAT IMAGE.','error')
        return redirect(url_for('change_credentials'))
    filename=f"account-{acct.id}-{secrets.token_hex(8)}.jpg"
    target=PROFILE_DIR/filename
    target.write_bytes(encoded.tobytes())
    if acct.profile_picture_path:
        old=BASE_DIR/'static'/acct.profile_picture_path.replace('static/','',1) if acct.profile_picture_path.startswith('static/') else BASE_DIR/'static'/acct.profile_picture_path
        try:
            if old.exists() and old != target: old.unlink()
        except OSError:
            pass
    acct.profile_picture_path=f"uploads/profiles/{filename}"
    log_audit('profile_picture_updated','Account',acct.id,{'filename':filename})
    db.session.commit()
    flash('PROFILE PICTURE UPDATED.','success')
    return redirect(url_for('change_credentials'))

@app.post('/account/profile-picture/remove')
@login_required()
def remove_profile_picture():
    acct=current_account()
    if acct.profile_picture_path:
        old=BASE_DIR/'static'/acct.profile_picture_path
        try:
            if old.exists(): old.unlink()
        except OSError:
            pass
        acct.profile_picture_path=None
        log_audit('profile_picture_removed','Account',acct.id)
        db.session.commit()
    flash('PROFILE PICTURE REMOVED.','success')
    return redirect(url_for('change_credentials'))

@app.route('/account', methods=['GET','POST'])
@login_required()
def change_credentials():
    acct=current_account()
    if request.method=='POST':
        current=request.form.get('current_password',''); new_user=request.form.get('username','').strip(); new_password=request.form.get('new_password',''); confirm=request.form.get('confirm_password','')
        if not new_user: return render_template('account.html',error='Username is required.',account=acct,first_time=acct.must_change_password)
        if not check_password_hash(acct.password_hash,current): return render_template('account.html',error='Current password is incorrect.',account=acct,first_time=acct.must_change_password)
        if len(new_password)<8 or new_password!=confirm: return render_template('account.html',error='Use an 8+ character password and make sure both new passwords match.',account=acct,first_time=acct.must_change_password)
        if new_user!=acct.username and Account.query.filter(func.lower(Account.username)==new_user.lower()).first(): return render_template('account.html',error='Username already exists.',account=acct,first_time=acct.must_change_password)
        acct.username=new_user; acct.password_hash=generate_password_hash(new_password); acct.must_change_password=False; log_audit('credentials_changed','Account',acct.id); db.session.commit(); return redirect(url_for('dashboard'))
    return render_template('account.html',account=acct,first_time=acct.must_change_password)

@app.get('/dashboard')
@login_required()
def dashboard():
    acct=current_account()
    if acct.role=='admin':
        try:
            maybe_create_fee_reminders()
        except Exception:
            db.session.rollback(); app.logger.exception('Fee reminder check failed')
    q=allowed_students_for_account(acct)
    student_count=q.count()
    start,end=school_year_bounds(); today=min(school_date(),end)
    students=student_order(q).limit(8).all()
    attendance_map = attendance_percentage_map([st.id for st in students], start, today)
    data=[{'name':st.name,'class_section':f'{st.class_name}-{st.section}' if st.section else st.class_name,
           'attendance':attendance_map.get(st.id,0)} for st in students]
    year_total=round(sum(x['attendance'] for x in data)/len(data),2) if data else 0

    audience_map={'teacher':'teachers','student':'students','parent':'parents'}
    if acct.role=='admin':
        aq=Announcement.query.filter(Announcement.published.is_(True), Announcement.audience.in_(['public','all','admin','teachers','students','parents']))
    elif acct.role=='parent':
        parent=Parent.query.filter_by(account_id=acct.id).first()
        clauses=[Announcement.audience.in_(['public','all']), and_(Announcement.audience=='parents',Announcement.parent_id.is_(None))]
        if parent: clauses.append(Announcement.parent_id==parent.id)
        aq=Announcement.query.filter(Announcement.published.is_(True), or_(*clauses))
    else:
        aq=Announcement.query.filter(Announcement.published.is_(True), Announcement.audience.in_(['public','all',audience_map.get(acct.role)]))
    announcements=aq.order_by(Announcement.published_at.desc(),Announcement.id.desc()).limit(5).all()

    class_teacher='—'
    if acct.role=='student':
        st=Student.query.filter_by(account_id=acct.id).first()
        if st:
            ta=TeacherAssignment.query.filter_by(class_name=st.class_name,section=st.section).first()
            if ta:
                t=db.session.get(Teacher,ta.teacher_id); class_teacher=t.name if t else '—'
    children_count=student_count if acct.role == 'parent' else None
    return render_template('dashboard.html',account=acct,role=acct.role,student_count=student_count,
        teacher_count=Teacher.query.filter_by(active=True).count() if acct.role=='admin' else None,
        attendance=data,year_attendance=year_total,announcements=announcements,
        working_today=is_working_day(school_date()),class_teacher=class_teacher,children_count=children_count)

@app.get('/admin')
@admin_required
def admin_home(): return redirect(url_for('dashboard'))

@app.get('/admin/register')
@admin_required
def admin_register_page():
    return render_template('register_student.html', class_options=class_list(), form={})

@app.post('/admin/register')
@admin_required
def admin_register_save():
    d=request.form; name=normalize_school_name(d.get('name')); adm=str(d.get('admission_number','')).strip().upper(); roll=str(d.get('roll_number','')).strip().upper()
    cls=normalize_class(d.get('class_name')); sec=normalize_section(d.get('section')); user=str(d.get('username','')).strip(); password=d.get('password','')
    second=(d.get('second_language') or '').strip().lower() or None; third=(d.get('third_language') or '').strip().lower() or None
    class_no = class_number(cls)
    if class_no>=9: third=None
    allowed_languages={'telugu','hindi','sanskrit'}
    if second and second not in allowed_languages:
        flash('SELECT A VALID 2ND LANGUAGE.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    if third and third not in allowed_languages:
        flash('SELECT A VALID 3RD LANGUAGE.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    if second and third and second == third:
        flash('2ND AND 3RD LANGUAGE MUST BE DIFFERENT.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    section_required=class_no not in {11,12}
    if not all([name,adm,roll,cls,user]) or (section_required and not sec) or len(password)<8:
        flash('NAME, ADMISSION NUMBER, ROLL NUMBER, CLASS, USERNAME AND AN 8+ CHARACTER TEMPORARY PASSWORD ARE REQUIRED. SECTION IS REQUIRED EXCEPT FOR XI/XII.','error')
        return render_template('register_student.html',class_options=class_list(),form=d)
    if class_number(cls) in range(5,9) and (not second or not third):
        flash('2ND AND 3RD LANGUAGE ARE REQUIRED FOR CLASSES V–VIII.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    if class_number(cls) in range(9,11) and not second:
        flash('2ND LANGUAGE IS REQUIRED FOR CLASSES IX–X.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    if Student.query.filter_by(admission_number=adm).first():
        flash('ADMISSION NUMBER ALREADY EXISTS.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    existing_account=Account.query.filter(func.lower(Account.username)==user.lower()).first()
    if existing_account and existing_account.active:
        flash('USERNAME ALREADY EXISTS.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    if existing_account and not existing_account.active:
        release_inactive_account_username(user)
    session['pending_student_registration']={'name':name,'admission_number':adm,'roll_number':roll,'class_name':cls,'section':sec,'username':user,'password_hash':generate_password_hash(password),'second_language':second,'third_language':third}
    if not face_available():
        acct=Account(username=user,password_hash=generate_password_hash(password),role='student',display_name=name,must_change_password=True)
        db.session.add(acct); db.session.flush()
        st=Student(name=name,admission_number=adm,roll_number=roll,class_name=cls,section=sec,second_language=second,third_language=third,account_id=acct.id)
        db.session.add(st); log_audit('student_created','Student',st.id,{'class':cls,'section':sec,'username':user,'face_frames':0}); db.session.commit(); session.pop('pending_student_registration',None)
        flash(f'STUDENT CREATED. USERNAME: {acct.username}. FACE ID CAN BE ADDED LATER.','success'); return redirect(url_for('admin_students'))
    return redirect(url_for('register_student_face'))

@app.get('/admin/register/face')
@admin_required
def register_student_face():
    pending=session.get('pending_student_registration')
    if not pending:
        flash('START STUDENT REGISTRATION FIRST.','error')
        return redirect(url_for('admin_register_page'))
    return render_template('register_face_capture.html', pending=pending)

@app.post('/admin/register/face/complete')
@admin_required
def register_student_face_complete():
    pending=session.get('pending_student_registration')
    if not pending:
        return jsonify({'error':'Registration session expired. Start again.'}),403
    frames=request.form.getlist('frames')
    if len(frames)<8:
        return jsonify({'error':'FACE ID IS REQUIRED. CAPTURE AT LEAST 8 VALID FRAMES.'}),400
    encs=[]
    for raw in frames:
        try:
            encs.append(encode_frame(raw))
        except Exception:
            continue
    if len(encs)<8:
        return jsonify({'error':f'Only {len(encs)} valid face frames were detected. Capture at least 8 good frames.'}),400
    if Student.query.filter_by(admission_number=pending['admission_number']).first():
        session.pop('pending_student_registration',None); return jsonify({'error':'Admission number already exists.'}),409
    existing_account=Account.query.filter(func.lower(Account.username)==pending['username'].lower()).first()
    if existing_account and existing_account.active:
        session.pop('pending_student_registration',None); return jsonify({'error':'Username already exists.'}),409
    if existing_account and not existing_account.active:
        release_inactive_account_username(pending['username'])
    acct=Account(username=pending['username'],password_hash=pending['password_hash'],role='student',display_name=pending['name'],must_change_password=True)
    db.session.add(acct); db.session.flush()
    s=Student(name=pending['name'],admission_number=pending['admission_number'],roll_number=pending['roll_number'],class_name=pending['class_name'],section=pending['section'],second_language=pending['second_language'],third_language=pending['third_language'],account_id=acct.id,face_encoding_json=json.dumps(encs),face_trained=True)
    db.session.add(s); db.session.flush()
    log_audit('student_created','Student',s.id,{'class':s.class_name,'section':s.section,'username':acct.username,'face_frames':len(encs)})
    db.session.commit(); session.pop('pending_student_registration',None)
    flash(f'STUDENT CREATED WITH FACE ID. USERNAME: {acct.username}. FIRST LOGIN REQUIRES A CREDENTIAL CHANGE.','success')
    return jsonify({'ok':True,'redirect':url_for('admin_students')})

@app.get('/admin/students')
@admin_required
def admin_students():
    """Admin student register.

    Students are intentionally NOT loaded on the initial page. The admin must
    first choose a class/section or enter a search/roll/admission filter. This
    keeps the page fast on large school datasets while preserving the same UI.
    """
    day = school_date()
    search = ' '.join((request.args.get('q') or '').split())
    cls = request.args.get('class_name', '').strip()
    sec = request.args.get('section', '').strip()
    roll = request.args.get('roll', '').strip()
    adm = request.args.get('admission', '').strip()
    page = max(1, request.args.get('page', 1, type=int))
    per_page = 25

    classes = [
        r[0] for r in db.session.query(Student.class_name)
        .filter(Student.active.is_(True))
        .filter(Student.class_name.isnot(None))
        .distinct()
        .order_by(Student.class_name)
        .all()
    ]
    sections = [
        r[0] for r in db.session.query(Student.section)
        .filter(Student.active.is_(True))
        .filter(Student.section.isnot(None))
        .filter(Student.section != '')
        .distinct()
        .order_by(Student.section)
        .all()
    ]

    # Do not query/load student rows until the admin actually requests them.
    has_selection = bool(
        search or cls or sec or roll or adm
    )

    rows = []
    total = 0

    if has_selection:
        q = Student.query.filter(Student.active.is_(True))

        if search:
            like = f'%{search}%'
            q = q.filter(or_(
                Student.name.ilike(like),
                Student.admission_number.ilike(like),
                Student.roll_number.ilike(like),
                Student.class_name.ilike(like),
                Student.section.ilike(like),
            ))

        if cls:
            q = q.filter(Student.class_name.ilike(cls))
        if sec:
            q = q.filter(Student.section.ilike(sec))
        if roll:
            q = q.filter(Student.roll_number.ilike(f'%{roll}%'))
        if adm:
            q = q.filter(Student.admission_number.ilike(f'%{adm}%'))

        total = q.count()
        rows_students = (
            student_order(q)
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )

        ids = [x.id for x in rows_students]
        ats = (
            Attendance.query
            .filter(
                Attendance.date == day,
                Attendance.student_id.in_(ids)
            )
            .all()
            if ids else []
        )
        att_by = {a.student_id: a for a in ats}
        rows = [
            {'student': st, 'attendance': att_by.get(st.id)}
            for st in rows_students
        ]

    return render_template(
        'students.html',
        rows=rows,
        day=day,
        class_name=cls,
        section=sec,
        classes=classes,
        sections=sections,
        page=page,
        per_page=per_page,
        total=total,
        search=search,
        roll=roll,
        admission=adm,
        has_selection=has_selection,
    )

@app.route('/admin/teachers', methods=['GET','POST'])
@admin_required
def admin_teachers():
    if request.method=='POST':
        name=normalize_school_name(request.form.get('name'))
        username=request.form.get('username','').strip()
        password=request.form.get('password','')
        phone=request.form.get('phone','').strip()
        if not name or not username or len(password)<8:
            flash('NAME, USERNAME AND AN 8+ CHARACTER TEMPORARY PASSWORD ARE REQUIRED.','error')
        elif Account.query.filter(func.lower(Account.username)==username.lower()).first():
            flash('USERNAME ALREADY EXISTS.','error')
        else:
            acct=Account(username=username,password_hash=generate_password_hash(password),role='teacher',display_name=name,must_change_password=True)
            db.session.add(acct); db.session.flush(); t=Teacher(name=name,phone=phone,account_id=acct.id); db.session.add(t); db.session.flush(); log_audit('teacher_created','Teacher',t.id,{'username':username}); db.session.commit(); flash('TEACHER ACCOUNT CREATED.','success')
    search=' '.join((request.args.get('q') or '').split()); cls=request.args.get('class_name','').strip(); sec=request.args.get('section','').strip(); page=max(1,request.args.get('page',1,type=int)); per_page=25
    q=Teacher.query.filter_by(active=True)
    if search:
        like=f'%{search}%'; q=q.outerjoin(Account,Teacher.account_id==Account.id).filter(or_(Teacher.name.ilike(like),Account.username.ilike(like)))
    if cls or sec:
        q=q.join(TeacherAssignment,TeacherAssignment.teacher_id==Teacher.id)
        if cls: q=q.filter(TeacherAssignment.class_name.ilike(cls))
        if sec: q=q.filter(TeacherAssignment.section.ilike(sec))
        q=q.distinct()
    total=q.count(); teachers=q.order_by(Teacher.name).offset((page-1)*per_page).limit(per_page).all()
    assignment_counts={t.id:TeacherAssignment.query.filter_by(teacher_id=t.id).count() for t in teachers}
    class_options=[x[0] for x in db.session.query(TeacherAssignment.class_name).distinct().order_by(TeacherAssignment.class_name).all()]
    section_options=[x[0] for x in db.session.query(TeacherAssignment.section).filter(TeacherAssignment.section!='').distinct().order_by(TeacherAssignment.section).all()]
    return render_template('teachers.html',teachers=teachers,assignment_counts=assignment_counts,search=search,class_name=cls,section=sec,page=page,per_page=per_page,total=total,class_options=class_options,section_options=section_options)

@app.route('/admin/announcements', methods=['GET','POST'])
@admin_required
def announcements():
    if request.method=='POST':
        title=request.form.get('title','').strip(); message=request.form.get('message','').strip(); audience=request.form.get('audience','all').strip(); publish=request.form.get('publish')=='1'
        allowed={'public','all','admin','students','teachers','parents'}
        if not title or not message or audience not in allowed:
            flash('Title, message and one valid audience are required.','error')
        else:
            a=Announcement(title=title,message=message,audience=audience,created_by=current_account().username,published=publish)
            if publish: a.published_at=datetime.utcnow()
            db.session.add(a); db.session.flush(); log_audit('announcement_created','Announcement',a.id,{'audience':audience,'published':publish})
            db.session.commit(); flash('Announcement created for the selected audience.','success')
    return render_template('announcements.html',announcements=Announcement.query.order_by(Announcement.created_at.desc()).limit(200).all())

@app.post('/admin/announcements/<int:aid>/publish')
@admin_required
def publish_announcement(aid):
    a=db.session.get(Announcement,aid) or abort(404); a.published=True; a.published_at=datetime.utcnow(); log_audit('announcement_published','Announcement',aid); db.session.commit(); return redirect(url_for('announcements'))

@app.post('/admin/announcements/<int:aid>/delete')
@admin_required
def delete_announcement(aid):
    a=db.session.get(Announcement,aid) or abort(404)
    db.session.delete(a); log_audit('announcement_deleted','Announcement',aid,{'title':a.title}); db.session.commit()
    flash('Announcement deleted.','success')
    return redirect(url_for('announcements'))

@app.get('/teacher')
@staff_required
def teacher_home(): return redirect(url_for('dashboard'))

@app.get('/students')
@login_required()
def students_self():
    acct=current_account(); rows=allowed_students_for_account(acct).all(); return render_template('portal_students.html',students=rows,role=acct.role)

@app.get('/api/school-day')
@staff_required
def school_day_api():
    # Scanner status must never fail just because a calendar row is malformed.
    # A failed calendar lookup falls back to the normal Mon-Fri rule.
    try:
        today=school_date()
        override=SchoolCalendar.query.filter_by(date=today).first()
        working=bool(override.is_working) if override is not None else weekly_default_is_working(today)
        reason=override.reason if override else None
        source='calendar' if override is not None else 'weekly-default'
    except Exception:
        db.session.rollback()
        today=school_date(); working=weekly_default_is_working(today); reason='Calendar unavailable; using weekday fallback'; source='fallback'
    effective=school_time()
    return jsonify({'date':today.isoformat(),'is_working':working,'reason':reason,'calendar_source':source,'override':source=='calendar','time':effective.strftime('%H:%M:%S'),'live_server_time':now_local().strftime('%H:%M'),'using_override':bool(get_school_clock_override()),'attendance_from':os.getenv('ATTENDANCE_PRESENT_FROM','07:30'),'late_after':os.getenv('ATTENDANCE_LATE_AFTER','08:30'),'absent_after':os.getenv('ATTENDANCE_ABSENT_AFTER','09:00')})

@app.get('/attendance/scan')
@staff_required
def attendance_scan_page():
    if not ip_allowed('SCANNER_ALLOWED_IPS'):
        abort(403)
    return render_template('attendance_scan.html',face_available=face_available(),scanner_ip_restricted=bool(os.getenv('SCANNER_ALLOWED_IPS','').strip()))

@app.get('/attendance')
@staff_required
def attendance_page():
    try: day=datetime.strptime(request.args.get('date',school_date().isoformat()),'%Y-%m-%d').date()
    except ValueError: day=school_date()
    q=allowed_students_for_account(current_account()); cls=normalize_class(request.args.get('class_name','')); sec=normalize_section(request.args.get('section',''))
    if cls: q=q.filter(Student.class_name==cls)
    if sec: q=q.filter(Student.section==sec)
    search=request.args.get('q','').strip()
    if search:
        like=f'%{search}%'; q=q.filter(or_(Student.name.ilike(like),Student.admission_number.ilike(like)))
    total=q.count(); page=max(1,request.args.get('page',1,type=int)); per_page=100
    students=student_order(q).offset((page-1)*per_page).limit(per_page).all()
    ids=[x.id for x in students]; ats=Attendance.query.filter(Attendance.date==day,Attendance.student_id.in_(ids)).all() if ids else []
    att_by={a.student_id:a for a in ats}; rows=[{'student':st,'attendance':att_by.get(st.id)} for st in students]
    classes=sorted({x.class_name for x in allowed_students_for_account(current_account()).with_entities(Student.class_name).distinct().all()},key=lambda x:(class_number(x) or 99,x))
    return render_template('attendance.html',rows=rows,day=day,face_available=face_available(),classes=classes,class_name=cls,section=sec,page=page,total=total,per_page=per_page,search=search)

@app.post('/api/attendance/mark')
@staff_required
def mark_attendance():
    d=request.get_json(silent=True) or {}; status=str(d.get('status','present')).lower()
    if status not in {'present','late','absent'}: return jsonify({'error':'Status must be present, late or absent.'}),400
    try: sid=int(d.get('student_id')); day=datetime.strptime(d.get('date',school_date().isoformat()),'%Y-%m-%d').date()
    except (TypeError,ValueError): return jsonify({'error':'Student and date are required.'}),400
    s=Student.query.filter_by(id=sid,active=True).first(); acct=current_account()
    if not s: abort(404)
    if acct.role!='admin' and allowed_students_for_account(acct).filter_by(id=sid).first() is None: abort(403)
    if day > school_date():
        return jsonify({'error':'Attendance cannot be marked for a future date.'}),400
    a=Attendance.query.filter_by(student_id=sid,date=day).first() or Attendance(student_id=sid,date=day)
    a.status=status
    a.time_in=school_time() if status in {'present','late'} and day == school_date() else None
    a.source='manual'; a.marked_by=acct.username; a.note=(d.get('note') or '').strip()[:500]; db.session.add(a)
    log_audit('attendance_marked','Student',sid,{'status':status,'date':str(day),'source':'manual'})
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        current = Attendance.query.filter_by(student_id=sid,date=day).first()
        return jsonify({'ok':True,'status':str(current.status).lower() if current else status,'already_exists':True}),200
    return jsonify({'ok':True,'status':status})

@app.post('/api/attendance/scan')
@staff_required
def attendance_scan():
    if not ip_allowed('SCANNER_ALLOWED_IPS'):
        return jsonify({'error':'Scanner access is restricted to the configured scanner network.'}),403
    if not face_available():
        return jsonify({'error':'Face recognition dependencies are not available on this server.'}),503
    payload=request.get_json(silent=True) or {}
    images=payload.get('images') or ([payload.get('image')] if payload.get('image') else [])
    images=[x for x in images[:7] if x]
    if not images:
        return jsonify({'error':'No camera frames received.'}),400
    day=school_date()
    try:
        working=is_working_day(day)
    except Exception:
        db.session.rollback(); working=weekly_default_is_working(day)
    if not working:
        return jsonify({'error':'Today is a non-working school day. Attendance cannot be marked.','school_day':False}),409
    now=school_time()
    try:
        present_from=dt_time.fromisoformat(os.getenv('ATTENDANCE_PRESENT_FROM','07:30'))
        late_after=dt_time.fromisoformat(os.getenv('ATTENDANCE_LATE_AFTER','08:30'))
        absent_after=dt_time.fromisoformat(os.getenv('ATTENDANCE_ABSENT_AFTER','09:00'))
    except ValueError:
        present_from,late_after,absent_after=dt_time(7,30),dt_time(8,30),dt_time(9,0)
    if now < present_from:
        return jsonify({'error':f'Attendance opens at {present_from.strftime("%H:%M")}.','school_day':True}),409
    if now >= absent_after:
        return jsonify({'error':f'The attendance window closed at {absent_after.strftime("%H:%M")}. An admin can manually mark the student present or late.','school_day':True}),409
    scan_status='late' if now >= late_after else 'present'
    tolerance=float(os.getenv('FACE_RECOGNITION_TOLERANCE','0.48'))
    acct=current_account()
    q=Student.query.filter_by(active=True,face_trained=True)
    if acct.role!='admin':
        q=allowed_students_for_account(acct).filter(Student.face_trained.is_(True))
    students=student_order(q).all()
    known=[]
    for st in students:
        if not st.face_encoding_json:
            continue
        try:
            encs=json.loads(st.face_encoding_json)
            if encs:
                known.append((st,encs))
        except Exception:
            continue
    if not known:
        return jsonify({'error':'No trained students are available for this scanner scope.'}),400
    votes,details={},{}
    valid_frames=0
    try:
        for image in images:
            frame=decode_data_url(image)
            locations,encs=recognize_faces(frame,upsample=1)
            if not encs:
                continue
            if len(encs)!=1:
                return jsonify({'error':'Only one person may be in front of the scanner'}),400
            ok,msg=image_quality(frame,locations[0])
            if not ok:
                continue
            valid_frames += 1
            student,score,second=best_match_for_encoding(encs[0],known,tolerance=tolerance)
            if student is not None and second is not None:
                margin_required=float(os.getenv('FACE_RECOGNITION_MIN_MARGIN','0.035'))
                if float(second) - float(score) < margin_required:
                    continue
            if student is not None:
                votes[student.id]=votes.get(student.id,0)+1
                details[student.id]=(student,score,second)
        if not votes:
            return jsonify({'error':'No confident match. Improve lighting, face the camera, and try again.','frames_used':valid_frames}),401
        winner_id,vote_count=max(votes.items(),key=lambda pair: pair[1])
        required_votes=1 if len(images)==1 else max(2,int(np.ceil(max(valid_frames,1)*0.60)))
        if vote_count<required_votes:
            return jsonify({'error':'Recognition was inconsistent. Hold still and scan again.','votes':vote_count,'required_votes':required_votes,'frames_used':valid_frames}),401
        student,score,_=details[winner_id]
        existing=Attendance.query.filter_by(student_id=student.id,date=day).with_for_update().first()
        if existing:
            status=str(existing.status or '').lower()
            if existing.source=='manual' and status=='absent':
                return jsonify({'error':f'{student.name} was manually marked absent by an admin. An admin must correct the record before face attendance can be applied.'}),409
            if status in {'present','late'}:
                return jsonify({'message':f'Attendance already marked for {student.name} ({status.title()})','student_name':student.name,'already_marked':True,'attendance_status':status,'votes':vote_count,'frames_used':valid_frames}),200
        if existing is None:
            existing=Attendance(student_id=student.id,date=day)
            db.session.add(existing)
        existing.status=scan_status
        existing.time_in=now.replace(microsecond=0)
        existing.source='face'
        existing.marked_by='Attendance Scanner'
        existing.note=None
        log_audit('face_attendance','Student',student.id,{'status':scan_status,'distance':round(float(score),4),'votes':vote_count,'frames_used':valid_frames})
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            existing2=Attendance.query.filter_by(student_id=student.id,date=day).first()
            if existing2:
                return jsonify({'message':f'Attendance already marked for {student.name}','student_name':student.name,'already_marked':True}),200
            raise
        return jsonify({'message':f'Welcome, {student.name}! Attendance marked as {scan_status.title()}.','student_id':student.id,'student_name':student.name,'attendance_status':scan_status,'school_time':now.strftime('%H:%M'),'distance':round(float(score),4),'votes':vote_count,'frames_used':valid_frames,'already_marked':False}),200
    except ValueError as exc:
        db.session.rollback()
        return json_error(str(exc))
    except Exception:
        db.session.rollback(); app.logger.exception('Attendance scan failed')
        return jsonify({'error':'Attendance scan failed. Please try again.'}),500

@app.get('/api/attendance/marklist')
@staff_required
def attendance_marklist():
    day=school_date()
    try:
        query=allowed_students_for_account(current_account()) if current_account().role!='admin' else Student.query.filter_by(active=True)
        students=student_order(query).all()
        ids=[st.id for st in students]
        records=Attendance.query.filter(Attendance.date==day,Attendance.student_id.in_(ids)).all() if ids else []
        by={r.student_id:r for r in records}
        counts={'present':0,'late':0,'absent':0,'not_marked':0}
        rows=[]
        for st in students:
            rec=by.get(st.id); status=str(rec.status).lower() if rec else 'not_marked'; counts[status]=counts.get(status,0)+1
            rows.append({'student_id':st.id,'name':st.name,'admission_number':st.admission_number,'class_name':st.class_name,'section':st.section or '', 'roll_number':st.roll_number or '', 'status':status,'time_in':rec.time_in.strftime('%H:%M:%S') if rec and rec.time_in else None,'source':rec.source if rec else None})
        return jsonify({'date':day.isoformat(),'school_time':school_time().strftime('%H:%M:%S'),'counts':counts,'total':len(rows),'rows':rows})
    except Exception:
        db.session.rollback(); app.logger.exception('Attendance marklist failed'); return jsonify({'error':'Could not load attendance marklist.'}),500

# Backward-compatible single-frame endpoint for older clients; the new scanner uses /api/attendance/scan.
@app.post('/api/attendance/recognize')
@staff_required
def recognize():
    payload=request.get_json(silent=True) or {}
    return attendance_scan.__wrapped__() if False else attendance_scan()

@app.route('/admin/students/<int:sid>/face', methods=['GET','POST'])
@admin_required
def student_face(sid):
    s=db.session.get(Student,sid) or abort(404)
    if request.method=='POST':
        frames=request.form.getlist('frames'); encs=[]
        for f in frames:
            try: encs.append(encode_frame(f))
            except Exception: pass
        if not encs: return render_template('face_capture.html',student=s,error='No valid face frames received.')
        s.face_encoding_json=json.dumps(encs); s.face_trained=True; log_audit('face_model_updated','Student',sid,{'frames':len(encs)}); db.session.commit(); flash('Face model updated.','success'); return redirect(url_for('admin_students'))
    return render_template('face_capture.html',student=s)

@app.get('/academics')
@staff_required
def academics():
    acct=current_account(); allowed=allowed_students_for_account(acct)
    subject_param=request.args.get('subject','').strip()
    # Teachers are class teachers only: a class teacher can enter marks for
    # every subject of the students in their assigned class/section.
    # Subject-level teacher assignment is intentionally not used.
    subjects_q=Subject.query.order_by(Subject.name)
    subjects=subjects_q.all()
    if subject_param and any(x.code==subject_param for x in subjects):
        subjects=[x for x in subjects if x.code==subject_param]
    students=student_order(allowed).limit(500).all()
    exams=Exam.query.order_by(Exam.order_index).all()
    return render_template('academics.html',exams=exams,subjects=subjects,all_subjects=subjects_q.all() if subject_param else subjects,students=students,selected_subject=subject_param)

def teacher_can_edit_mark(teacher, student, subject):
    if not teacher or not teacher.active:
        return False
    class_assigned = TeacherAssignment.query.filter_by(
        teacher_id=teacher.id, class_name=student.class_name, section=student.section
    ).first() is not None
    if class_assigned:
        return True
    return TeacherSubjectAssignment.query.filter_by(
        teacher_id=teacher.id, subject_code=subject.code
    ).first() is not None

@app.post('/api/marks')
@login_required('teacher')
def save_mark():
    d=request.get_json() or {}
    try:
        sid=int(d['student_id']); exam=db.session.get(Exam,int(d['exam_id']))
    except (KeyError,TypeError,ValueError):
        return jsonify({'error':'Invalid student or exam.'}),400
    subj=Subject.query.filter_by(code=d.get('subject_code','')).first()
    val=float(d['marks']) if d.get('marks') not in ('',None) else None
    student=allowed_students_for_account(current_account()).filter_by(id=sid).first()
    if not student or not exam or not subj: abort(403)
    teacher=Teacher.query.filter_by(account_id=current_account().id,active=True).first()
    if not teacher_can_edit_mark(teacher, student, subj):
        return jsonify({'error':'You are not assigned to enter this subject for this class/section.'}),403
    if val is not None and (val<0 or val>exam.max_marks): return jsonify({'error':f'Marks must be between 0 and {exam.max_marks}.'}),400
    m=Mark.query.filter_by(student_id=sid,subject_code=subj.code,exam_id=exam.id).first() or Mark(student_id=sid,subject_code=subj.code,exam_id=exam.id,max_marks=exam.max_marks)
    if m.locked: return jsonify({'error':'This mark is locked.'}),409
    m.marks=val; m.updated_by=current_account().username; m.max_marks=exam.max_marks; db.session.add(m); log_audit('mark_updated','Mark',m.id,{'student_id':sid,'exam':exam.name,'subject':subj.name,'marks':val}); db.session.commit(); return jsonify({'ok':True})

@app.post('/academics/import')
@staff_required
def academics_import():
    upload=request.files.get('file')
    if not upload or not upload.filename.lower().endswith(('.xlsx','.xlsm')):
        flash('Upload an Excel workbook (.xlsx/.xlsm).','error'); return redirect(url_for('academics'))
    try:
        from openpyxl import load_workbook
        wb=load_workbook(upload,read_only=True,data_only=True)
        ws=wb.active
        rows=list(ws.iter_rows(values_only=True))
        if len(rows)<2: raise ValueError('The workbook is empty.')
        headers=[str(x or '').strip().lower().replace(' ','_') for x in rows[0]]
        aliases={
            'admission':'admission','admission_number':'admission','student_admission_number':'admission',
            'exam':'exam','exam_name':'exam','subject':'subject','subject_name':'subject',
            'marks':'marks','mark':'marks','score':'marks'
        }
        normalized=[aliases.get(h,h) for h in headers]
        required={'admission','exam','subject','marks'}
        if not required.issubset(set(normalized)):
            raise ValueError('Required columns: Admission, Exam, Subject, Marks')
        idx={h:i for i,h in enumerate(normalized)}
        students_by_admission={s.admission_number.strip().upper():s for s in Student.query.filter_by(active=True).all()}
        exams_by_name={e.name.strip().lower():e for e in Exam.query.filter_by(active=True).all()}
        subjects_by_name={s.name.strip().lower():s for s in Subject.query.filter_by(active=True).all()}
        acct=current_account(); teacher=Teacher.query.filter_by(account_id=acct.id,active=True).first() if acct.role=='teacher' else None
        class_pairs=set((a.class_name,a.section) for a in TeacherAssignment.query.filter_by(teacher_id=teacher.id).all()) if teacher else set()
        teacher_subjects=set(a.subject_code for a in TeacherSubjectAssignment.query.filter_by(teacher_id=teacher.id).all()) if teacher else set()
        changed=0; skipped=0
        for raw_row in rows[1:]:
            admission=str(raw_row[idx['admission']] or '').strip().upper(); exam_name=str(raw_row[idx['exam']] or '').strip().lower(); subject_name=str(raw_row[idx['subject']] or '').strip().lower(); raw=raw_row[idx['marks']]
            if not admission or not exam_name or not subject_name or raw in (None,''):
                skipped += 1; continue
            student=students_by_admission.get(admission); exam=exams_by_name.get(exam_name); subject=subjects_by_name.get(subject_name)
            if not student or not exam or not subject:
                skipped += 1; continue
            if acct.role=='teacher' and not (
                (student.class_name,student.section) in class_pairs or subject.code in teacher_subjects
            ):
                skipped += 1; continue
            try:
                value=float(raw)
            except (TypeError,ValueError):
                skipped += 1; continue
            if value<0 or value>exam.max_marks:
                skipped += 1; continue
            mark=Mark.query.filter_by(student_id=student.id,subject_code=subject.code,exam_id=exam.id).first()
            if mark is None:
                mark=Mark(student_id=student.id,subject_code=subject.code,exam_id=exam.id,max_marks=exam.max_marks)
                db.session.add(mark)
            if mark.locked:
                skipped += 1; continue
            mark.marks=value; mark.max_marks=exam.max_marks; mark.updated_by=acct.username; changed += 1
        log_audit('marks_excel_import','Academics',extra={'updated':changed,'skipped':skipped,'filename':upload.filename}); db.session.commit()
        flash(f'Imported {changed} mark(s). Skipped {skipped} invalid, unknown, unauthorized or locked row(s).','success')
    except Exception as exc:
        db.session.rollback(); app.logger.exception('Excel marks import failed'); flash(f'Could not import workbook: {exc}','error')
    return redirect(url_for('academics'))

@app.get('/academics/import-template.xlsx')
@staff_required
def academics_import_template():
    rows=[
        {'Admission':'STUDENT-001','Exam':'PT-1','Subject':'Mathematics','Marks':35},
        {'Admission':'STUDENT-001','Exam':'Final Examination','Subject':'Mathematics','Marks':72},
    ]
    buf=build_xlsx(rows,'Marks Import')
    return send_file(buf,as_attachment=True,download_name='marks_import_template.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.context_processor
def template_helpers():
    acct = current_account()
    return {
        'subjects_for_class': subjects_for_class,
        'school_date': school_date,
        'school_time': school_time,
        'school_now': school_now,
        'clock_override_time': (db.session.get(SchoolClock,1).override_time if (db.session.get(SchoolClock,1) is not None) else None),
        'grade_for_percent': grade_for_percent,
        'current_account': current_account,
        'profile_picture_url': profile_picture_url,
        'me': acct,
        'csrf_token': session.get('csrf',''),
    }

@app.get('/results')
@login_required()
def results():
    acct=current_account(); students=student_order(allowed_students_for_account(acct)).limit(1000).all(); exams=Exam.query.order_by(Exam.order_index).all(); ids=[s.id for s in students]
    marks=Mark.query.filter(Mark.student_id.in_(ids)).all() if ids else []
    by_student={}
    for m in marks: by_student.setdefault(m.student_id,[]).append(m)
    rows=[]
    for s in students:
        total=0; max_total=0; by_exam=[]; smarks=by_student.get(s.id,[])
        for e in exams:
            ms=[m for m in smarks if m.exam_id==e.id]; got=sum(m.marks or 0 for m in ms); mx=sum(m.max_marks for m in ms); pct=round(got/mx*100,2) if mx else 0
            total += got; max_total += mx; by_exam.append((e.name,got,mx,pct,grade_for_percent(pct)))
        pct=round(total/max_total*100,2) if max_total else 0
        rows.append({'student':s,'by_exam':by_exam,'percentage':pct,'grade':grade_for_percent(pct),'result':pass_fail_for_percent(pct)})
    setting=SchoolSetting.query.filter_by(key='academic_session').first(); school_session=setting.value if setting else os.getenv('ACADEMIC_SESSION','2026-27')
    return render_template('results.html',rows=rows,exams=exams,school_session=school_session)

@app.get('/results/export.xlsx')
@login_required()
def results_xlsx():
    rows=[]
    for s in allowed_students_for_account(current_account()).all():
        for e in Exam.query.order_by(Exam.order_index).all():
            ms=Mark.query.filter_by(student_id=s.id,exam_id=e.id).all(); rows.append({'Admission':s.admission_number,'Name':s.name,'Class':s.class_name,'Section':s.section,'Exam':e.name,'Marks':sum(m.marks or 0 for m in ms),'Max':sum(m.max_marks for m in ms)})
    return send_file(build_xlsx(rows,'Results'),as_attachment=True,download_name='results.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/admin/attendance/export.xlsx')
@staff_required
def attendance_xlsx():
    try:
        day=datetime.strptime(request.args.get('date',school_date().isoformat()),'%Y-%m-%d').date()
    except ValueError:
        day=school_date()
    cls=normalize_class(request.args.get('class_name','')); sec=normalize_section(request.args.get('section',''))
    q=allowed_students_for_account(current_account())
    if cls: q=q.filter(Student.class_name==cls)
    if sec: q=q.filter(Student.section==sec)
    students=student_order(q).all()
    ids=[s.id for s in students]
    ats=Attendance.query.filter(Attendance.date==day,Attendance.student_id.in_(ids)).all() if ids else []
    by={a.student_id:a for a in ats}
    rows=[]
    for s in students:
        a=by.get(s.id)
        rows.append({'Roll':s.roll_number or '', 'Admission':s.admission_number, 'Name':s.name, 'Class':s.class_name, 'Section':s.section or '', 'Status':(a.status if a else 'absent').title(), 'Time':(a.time_in.strftime('%I:%M %p') if a and a.time_in else '')})
    return send_file(build_xlsx(rows,'Attendance Register'),as_attachment=True,download_name=f'attendance_{day.isoformat()}.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/admin/attendance/export.pdf')
@staff_required
def attendance_pdf():
    try:
        day=datetime.strptime(request.args.get('date',school_date().isoformat()),'%Y-%m-%d').date()
    except ValueError:
        day=school_date()
    cls=normalize_class(request.args.get('class_name','')); sec=normalize_section(request.args.get('section',''))
    q=allowed_students_for_account(current_account())
    if cls: q=q.filter(Student.class_name==cls)
    if sec: q=q.filter(Student.section==sec)
    students=student_order(q).all()
    ids=[s.id for s in students]
    ats=Attendance.query.filter(Attendance.date==day,Attendance.student_id.in_(ids)).all() if ids else []
    by={a.student_id:a for a in ats}
    rows=[]
    for s in students:
        a=by.get(s.id)
        rows.append({'Roll':s.roll_number or '', 'Admission':s.admission_number, 'Name':s.name, 'Class':s.class_name, 'Section':s.section or '', 'Status':(a.status if a else 'absent').title(), 'Time':(a.time_in.strftime('%I:%M %p') if a and a.time_in else '')})
    return send_file(build_pdf(rows,'DAV PS KKP • Attendance Register',subtitle=day.strftime('%d.%m.%Y')),as_attachment=True,download_name=f'attendance_{day.isoformat()}.pdf',mimetype='application/pdf')


@app.get('/results/export.pdf')
@login_required()
def results_pdf():
    rows=[]
    for s in allowed_students_for_account(current_account()).all():
        for e in Exam.query.order_by(Exam.order_index).all():
            ms=Mark.query.filter_by(student_id=s.id,exam_id=e.id).all(); rows.append({'Admission':s.admission_number,'Name':s.name,'Class':s.class_name,'Section':s.section,'Exam':e.name,'Marks':sum(m.marks or 0 for m in ms),'Max':sum(m.max_marks for m in ms)})
    return send_file(build_pdf(rows,'Results'),as_attachment=True,download_name='results.pdf',mimetype='application/pdf')

def student_can_see_results(student):
    final=Exam.query.filter_by(is_final=True).first()
    if not final: return False
    pub=ResultPublication.query.filter_by(exam_id=final.id,class_name=student.class_name,section=student.section,published=True).first()
    return bool(pub)

@app.get('/my/attendance')
@login_required()
def my_attendance():
    acct=current_account(); rows=[]
    for s in allowed_students_for_account(acct).all():
        start,end=school_year_bounds(); today=min(school_date(),end)
        y=attendance_summary(s.id,start,today)
        term1_end=date(start.year,9,30); term2_start=date(start.year,10,1)
        t1=attendance_summary(s.id,start,min(today,term1_end))
        t2=attendance_summary(s.id,term2_start,today) if today>=term2_start else {'working_days':0,'present':0,'absent':0,'percentage':0}
        rows.append({'student':s,'year':y,'term1':t1,'term2':t2})
    return render_template('my_attendance.html',rows=rows)

@app.get('/my/results')
@login_required()
def my_results():
    acct=current_account(); data=[]
    for s in allowed_students_for_account(acct).all():
        data.append((s,student_can_see_results(s)))
    return render_template('my_results.html',rows=data)

@app.get('/teacher/signature')
@staff_required
def teacher_signature():
    t=Teacher.query.filter_by(account_id=current_account().id).first()
    if not t: abort(404)
    return render_template('teacher_signature.html',teacher=t)

@app.post('/teacher/signature')
@staff_required
def teacher_signature_save():
    t=Teacher.query.filter_by(account_id=current_account().id).first() or abort(404)

    # Teachers can either draw a signature on the canvas or upload an image.
    uploaded = request.files.get('signature_file')
    uploaded_value = ''
    if uploaded and uploaded.filename:
        raw = uploaded.read()
        if len(raw) > 2 * 1024 * 1024:
            flash('SIGNATURE IMAGE MUST BE 2 MB OR SMALLER.','error')
            return redirect(url_for('teacher_signature'))
        ext = Path(uploaded.filename).suffix.lower()
        mimetype = (uploaded.mimetype or '').lower()
        allowed = {'.png':'image/png', '.jpg':'image/jpeg', '.jpeg':'image/jpeg', '.webp':'image/webp'}
        if ext not in allowed or mimetype not in set(allowed.values()):
            flash('UPLOAD A PNG, JPG, JPEG, OR WEBP SIGNATURE IMAGE.','error')
            return redirect(url_for('teacher_signature'))
        uploaded_value = f"data:{allowed[ext]};base64,{base64.b64encode(raw).decode('ascii')}"

    drawn_value = request.form.get('signature_data','').strip()
    if uploaded_value:
        t.signature_data = uploaded_value[:200000]
    elif drawn_value.startswith('data:image/'):
        t.signature_data = drawn_value[:200000]
    else:
        flash('DRAW OR UPLOAD A SIGNATURE FIRST.','error')
        return redirect(url_for('teacher_signature'))

    log_audit('teacher_signature_updated','Teacher',t.id)
    db.session.commit()
    flash('Signature saved for future report cards.','success')
    return redirect(url_for('teacher_signature'))

@app.get('/admin/assignments')
@admin_required
def assignments():
    assignments=TeacherAssignment.query.order_by(TeacherAssignment.class_name,TeacherAssignment.section).all()
    teachers=Teacher.query.filter_by(active=True).order_by(Teacher.name).all()
    teacher_names={t.id:t.name for t in teachers}
    return render_template('assignments.html',teachers=teachers,assignments=assignments,teacher_names=teacher_names,class_options=class_list())

@app.post('/admin/assignments')
@admin_required
def assignments_save():
    tid=int(request.form.get('teacher_id')); cls=normalize_class(request.form.get('class_name')); sec=normalize_section(request.form.get('section'))
    teacher=db.session.get(Teacher,tid)
    if not teacher or not cls:
        flash('SELECT A VALID TEACHER AND CLASS.','error'); return redirect(url_for('assignments'))
    if sec and class_number(cls) < 9 and len(sec)>2:
        flash('SECTION MUST BE A SHORT CLASS SECTION SUCH AS A OR B.','error'); return redirect(url_for('assignments'))
    existing=TeacherAssignment.query.filter_by(class_name=cls,section=sec).first()
    if existing:
        flash('THIS CLASS/SECTION ALREADY HAS A CLASS TEACHER.','error'); return redirect(url_for('assignments'))
    x=TeacherAssignment(teacher_id=tid,class_name=cls,section=sec)
    db.session.add(x)
    log_audit('teacher_assignment_created','TeacherAssignment',extra={'teacher_id':tid,'class_name':cls,'section':sec})
    db.session.commit()
    flash('CLASS TEACHER ASSIGNMENT ADDED.','success')
    return redirect(url_for('assignments'))

@app.get('/api/subjects-for-class')
@admin_required
def subjects_for_class_api():
    cls=normalize_class(request.args.get('class_name','')); second=request.args.get('second_language'); third=request.args.get('third_language')
    return jsonify([{'code':c,'name':n} for c,n in subject_options_for_class(cls,second,third) if not c.endswith('_placeholder')])


def _json_dict(value):
    try: return json.loads(value or '{}')
    except Exception: return {}


def _can_edit_report_card(acct, student):
    if acct.role == 'admin': return True
    if acct.role != 'teacher': return False
    teacher=Teacher.query.filter_by(account_id=acct.id,active=True).first()
    if not teacher: return False
    return TeacherAssignment.query.filter_by(teacher_id=teacher.id,class_name=student.class_name,section=student.section).first() is not None

@app.route('/report-card/config/<int:sid>', methods=['GET','POST'])
@staff_required
def report_card_config(sid):
    student=db.session.get(Student,sid) or abort(404); acct=current_account()
    if not _can_edit_report_card(acct, student): abort(403)
    setting=SchoolSetting.query.filter_by(key='academic_session').first(); session_name=setting.value if setting else os.getenv('ACADEMIC_SESSION','2026-27')
    cfg=ReportCardConfig.query.filter_by(student_id=sid).first()
    if not cfg:
        cfg=ReportCardConfig(student_id=sid,academic_session=session_name,next_academic_session=str(int(session_name[:4])+1)+'-'+str(int(session_name[:4])+2)[-2:] if session_name[:4].isdigit() else '', date_result=school_date().strftime('%d.%m.%Y'))
        ta=TeacherAssignment.query.filter_by(class_name=student.class_name,section=student.section).first()
        if ta:
            tt=Teacher.query.get(ta.teacher_id); cfg.class_teacher_name=tt.name if tt else ''
        db.session.add(cfg); db.session.flush()
    if not cfg.class_teacher_name:
        ta0=TeacherAssignment.query.filter_by(class_name=student.class_name,section=student.section).first()
        if ta0:
            tt0=Teacher.query.get(ta0.teacher_id); cfg.class_teacher_name=tt0.name if tt0 else ''
    if not cfg.date_result: cfg.date_result=school_date().strftime('%d.%m.%Y')
    if not cfg.next_academic_session and session_name[:4].isdigit(): cfg.next_academic_session=f'{int(session_name[:4])+1}-{str(int(session_name[:4])+2)[-2:]}'
    db.session.flush()
    subjects=subjects_for_class(student.class_name,student.second_language,student.third_language)
    if request.method=='POST':
        def clean_text(name,limit=4000): return str(request.form.get(name,'')).strip()[:limit]
        cfg.academic_session=clean_text('academic_session',30) or session_name
        cfg.house=clean_text('house',80); cfg.class_teacher_name=clean_text('class_teacher_name',160)
        cfg.remarks=clean_text('remarks'); cfg.principal_remarks=clean_text('principal_remarks'); cfg.date_result=clean_text('date_result',80)
        cfg.next_academic_session=clean_text('next_academic_session',30); cfg.session_begins=clean_text('session_begins',40); cfg.summer_break_from=clean_text('summer_break_from',40); cfg.school_reopens=clean_text('school_reopens',40)
        co={k:request.form.get(k,'').strip() for k in ['work_education','art_education','health_physical']}
        dis={k:request.form.get(k,'').strip() for k in ['discipline','regularity','punctuality']}
        health={k:request.form.get(k,'').strip() for k in ['term1_height','term1_weight','term2_height','term2_weight']}
        layout={k:request.form.get(k,'').strip()[:200] for k in ['report_title','attendance_title','scholastic_title','development_title','guide_title','class_details_title','pass_rule','teacher_signature_label','principal_signature_label','parent_signature_label']}
        cfg.co_scholastic_json=json.dumps(co); cfg.discipline_json=json.dumps(dis); cfg.health_json=json.dumps(health); cfg.layout_json=json.dumps(layout); cfg.updated_by=acct.username
        for subj in subjects:
            component=AssessmentComponent.query.filter_by(student_id=sid,subject_code=subj.code).first() or AssessmentComponent(student_id=sid,subject_code=subj.code)
            for field in ['multiple_assessment','subject_enrichment','portfolio']:
                raw=request.form.get(f'{subj.code}_{field}','')
                try: setattr(component,field,max(0,min(5,float(raw))) if raw.strip() else 0)
                except ValueError: setattr(component,field,0)
            # internal is computed from best PT + the three 5-point components
            db.session.add(component)
        log_audit('report_card_config_updated','ReportCardConfig',cfg.id,{'student_id':sid}); db.session.commit(); flash('REPORT CARD DETAILS SAVED.','success'); return redirect(url_for('results'))
    components={x.subject_code:x for x in AssessmentComponent.query.filter_by(student_id=sid).all()}
    layout=_json_dict(getattr(cfg,'layout_json','{}'))
    return render_template('report_card_config.html',student=student,cfg=cfg,co=_json_dict(cfg.co_scholastic_json),dis=_json_dict(cfg.discipline_json),health=_json_dict(cfg.health_json),layout=layout,subjects=subjects,components=components,session_name=session_name)

@app.get('/report-card/<int:sid>.pdf')
@login_required()
def report_card(sid):
    s=db.session.get(Student,sid); acct=current_account()
    if not s or allowed_students_for_account(acct).filter_by(id=sid).first() is None: abort(403)
    if acct.role in {'student','parent'} and not student_can_see_results(s):
        return render_template('locked_result.html',student=s), 403
    setting=SchoolSetting.query.filter_by(key='academic_session').first(); session_name=setting.value if setting else os.getenv('ACADEMIC_SESSION','2026-27')
    cfg=ReportCardConfig.query.filter_by(student_id=sid).first()
    config={}
    if cfg:
        config={'house':cfg.house,'class_teacher_name':cfg.class_teacher_name,'co_scholastic':_json_dict(cfg.co_scholastic_json),'discipline':_json_dict(cfg.discipline_json),'health':_json_dict(cfg.health_json),'remarks':cfg.remarks or '','principal_remarks':cfg.principal_remarks or '','date_result':cfg.date_result or '','next_academic_session':cfg.next_academic_session or '','session_begins':cfg.session_begins or '','summer_break_from':cfg.summer_break_from or '','school_reopens':cfg.school_reopens or '','layout':_json_dict(getattr(cfg,'layout_json','{}'))}
    start,end=school_year_bounds(); today=min(school_date(),end); t1_end=date(start.year,9,30); t2_start=date(start.year,10,1)
    att=attendance_summary(s.id,start,today); t1=attendance_summary(s.id,start,min(today,t1_end)); t2=attendance_summary(s.id,t2_start,today) if today>=t2_start else {'working_days':0,'present':0,'absent':0,'percentage':0}
    exams=Exam.query.order_by(Exam.order_index).all(); subjects=subjects_for_class(s.class_name,s.second_language,s.third_language)
    smarks=Mark.query.filter_by(student_id=s.id).all(); marks_by_subject={(m.subject_code,m.exam_id):m for m in smarks}
    comps={x.subject_code:{'multiple_assessment':x.multiple_assessment or 0,'subject_enrichment':x.subject_enrichment or 0,'portfolio':x.portfolio or 0} for x in AssessmentComponent.query.filter_by(student_id=s.id).all()}
    teacher_signature=None
    ta=TeacherAssignment.query.filter_by(class_name=s.class_name,section=s.section).first()
    if ta:
        teacher=Teacher.query.get(ta.teacher_id); teacher_signature=teacher.signature_data if teacher else None
    logo_path=BASE_DIR/'static'/'branding'/'dav-ps-kkp-logo.png'
    buf=build_report_card(s,session_name,att,t1,t2,subjects,exams,marks_by_subject,teacher_signature,config,logo_path,comps)
    return send_file(buf,as_attachment=True,download_name=f'report_card_{s.admission_number}.pdf',mimetype='application/pdf')

@app.route('/admin/accounts', methods=['GET','POST'])
@admin_required
def admin_accounts():
    if request.method=='POST':
        role=request.form.get('role'); username=request.form.get('username','').strip(); password=request.form.get('password',''); name=normalize_school_name(request.form.get('name'))
        child_admissions=[x.strip().upper() for x in request.form.get('child_admissions','').replace('\n',',').split(',') if x.strip()]
        child_ids=[s.id for s in Student.query.filter(Student.admission_number.in_(child_admissions)).all()]
        if role not in {'admin','teacher','parent'} or not username or len(password)<8 or not name:
            flash('Name, username, role and 8+ character temporary password are required. Student accounts are created only through registration.','error')
        elif role=='parent' and not child_ids:
            flash('A parent account must be linked to at least one student.','error')
        else:
            existing=Account.query.filter(func.lower(Account.username)==username.lower()).first()
            if existing and existing.active:
                flash('Username already exists.','error')
            else:
                if existing and not existing.active:
                    release_inactive_account_username(username)
                acct=Account(username=username,password_hash=generate_password_hash(password),role=role,display_name=name,must_change_password=True)
                db.session.add(acct); db.session.flush()
                if role=='parent':
                    parent=Parent.query.filter(Parent.account_id.is_(None), func.lower(Parent.name)==name.lower()).order_by(Parent.id.desc()).first()
                    if parent is None:
                        parent=Parent(name=name,account_id=acct.id,phone=request.form.get('phone','').strip(),email=request.form.get('email','').strip())
                        db.session.add(parent); db.session.flush()
                    else:
                        parent.account_id=acct.id; parent.name=name; parent.phone=request.form.get('phone','').strip(); parent.email=request.form.get('email','').strip()
                        db.session.query(ParentStudent).filter_by(parent_id=parent.id).delete(synchronize_session=False)
                    for sid in child_ids: db.session.add(ParentStudent(parent_id=parent.id,student_id=sid))
                elif role=='teacher':
                    db.session.add(Teacher(name=name,phone=request.form.get('phone','').strip(),email=request.form.get('email','').strip(),account_id=acct.id))
                log_audit('account_created','Account',acct.id,{'role':role,'children':child_ids}); db.session.commit(); flash(f'{role.title()} account created.','success')

    role=request.args.get('role','').lower(); search=request.args.get('q','').strip(); page=max(1,request.args.get('page',1,type=int)); per_page=25
    admin_items=[]; admin_count=0
    if not role or role=='admin':
        aq=Account.query.filter_by(role='admin',active=True)
        if search: aq=aq.filter(or_(Account.username.ilike(f'%{search}%'),Account.display_name.ilike(f'%{search}%')))
        admin_count=aq.count(); admin_items=aq.order_by(Account.display_name).offset((page-1)*per_page).limit(per_page).all()

    teacher_items=[]; teacher_count=0; assignment_map={}; teacher_accounts={}
    if not role or role=='teacher':
        tq=Teacher.query.filter_by(active=True)
        if search:
            tq=tq.outerjoin(Account,Teacher.account_id==Account.id).filter(or_(Teacher.name.ilike(f'%{search}%'),Account.username.ilike(f'%{search}%')))
        teacher_count=tq.count(); teacher_items=tq.order_by(Teacher.name).offset((page-1)*per_page).limit(per_page).all()
        tids=[t.id for t in teacher_items]
        for t in teacher_items: teacher_accounts[t.id]=t.account
        if tids:
            assigns=TeacherAssignment.query.filter(TeacherAssignment.teacher_id.in_(tids)).order_by(TeacherAssignment.class_name,TeacherAssignment.section).all()
            for a in assigns: assignment_map.setdefault(a.teacher_id,[]).append(f'{a.class_name}-{a.section}' if a.section else a.class_name)

    student_items=[]; student_count=0; parent_count_by_student={}
    if not role or role=='student':
        sq=Student.query.filter_by(active=True)
        if search:
            sq=sq.outerjoin(Account,Student.account_id==Account.id).filter(or_(Student.name.ilike(f'%{search}%'),Student.admission_number.ilike(f'%{search}%'),Student.roll_number.ilike(f'%{search}%'),Account.username.ilike(f'%{search}%')))
        student_count=sq.count(); student_items=student_order(sq).offset((page-1)*per_page).limit(per_page).all()
        sids=[st.id for st in student_items]
        if sids:
            links=ParentStudent.query.filter(ParentStudent.student_id.in_(sids)).all()
            for link in links: parent_count_by_student[link.student_id]=parent_count_by_student.get(link.student_id,0)+1

    parent_items=[]; parent_count=0; child_map={}
    if not role or role=='parent':
        pq=Parent.query.outerjoin(Account,Parent.account_id==Account.id).filter(or_(Account.id.is_(None),Account.active.is_(True)))
        if search:
            like=f'%{search}%'; pq=pq.filter(or_(Parent.name.ilike(like),Account.username.ilike(like),Parent.phone.ilike(like),Parent.email.ilike(like)))
        parent_count=pq.count(); parent_items=pq.order_by(Parent.name).offset((page-1)*per_page).limit(per_page).all()
        pids=[p.id for p in parent_items]
        if pids:
            links=ParentStudent.query.filter(ParentStudent.parent_id.in_(pids)).all()
            sids=[x.student_id for x in links]
            smap={st.id:st for st in Student.query.filter(Student.id.in_(sids)).all()} if sids else {}
            for link in links:
                st=smap.get(link.student_id)
                if st: child_map.setdefault(link.parent_id,[]).append(st)

    totals={'admin':admin_count,'teacher':teacher_count,'student':student_count,'parent':parent_count}
    return render_template('accounts.html',role=role,search=search,page=page,per_page=per_page,
        admin_items=admin_items,teacher_items=teacher_items,student_items=student_items,parent_items=parent_items,
        totals=totals,assignment_map=assignment_map,child_map=child_map,teacher_accounts=teacher_accounts,parent_count_by_student=parent_count_by_student)

FEE_CLASS_GROUPS = ['Nursery & LKG','UKG to IV','V to IX','X','XI & XII']
FEE_TERM_LABELS = [('term_i','TERM-I'),('term_ii','TERM-II'),('term_iii','TERM-III'),('term_iv','TERM-IV')]
DEFAULT_FEE_PAYMENT_WINDOWS_2026_27 = {
    'term_i': ('1st Term', date(2026,4,16), date(2026,6,25), date(2026,6,26)),
    'term_ii': ('2nd Term', date(2026,7,1), date(2026,9,25), date(2026,9,26)),
    'term_iii': ('3rd Term', date(2026,10,1), date(2026,12,25), date(2026,12,26)),
    'term_iv': ('4th Term', date(2027,1,1), date(2027,2,20), date(2027,2,21)),
}
DEFAULT_FEE_STRUCTURE_2026_27 = {
    'Nursery & LKG': (19700,19700,19700,0),
    'UKG to IV': (14775,14775,14775,14775),
    'V to IX': (15480,15480,15480,15480),
    'X': (16395,16395,16395,16395),
    'XI & XII': (17700,17700,17700,17700),
}

def fee_invoice_status_for_term(student_id, session_name, term_key):
    label = dict(FEE_TERM_LABELS).get(term_key, term_key)
    title = f'School Fee {session_name} • {label}'
    invoices = FeeInvoice.query.filter_by(student_id=student_id, title=title).order_by(FeeInvoice.created_at.desc()).all()
    if not invoices:
        return {'state':'unbilled','paid':0.0,'due':0.0,'balance':None,'invoice_id':None}
    inv = invoices[0]
    paid = sum(float(x.amount or 0) for x in inv.payments)
    due = float(inv.amount_due or 0)
    balance = max(0.0, due-paid)
    state = 'paid' if balance <= 0 else ('partial' if paid > 0 else 'unpaid')
    return {'state':state,'paid':paid,'due':due,'balance':balance,'invoice_id':inv.id}

def fee_group_for_student(student):
    raw=str(student.class_name or '').strip().upper()
    if raw in {'NURSERY','LKG'}: return 'Nursery & LKG'
    if raw=='UKG' or raw in {'1','2','3','4','I','II','III','IV'}: return 'UKG to IV'
    try:
        n=int(raw)
        if 1 <= n <= 4: return 'UKG to IV'
        if 5 <= n <= 9: return 'V to IX'
        if n==10: return 'X'
        if n in (11,12): return 'XI & XII'
    except Exception: pass
    if raw in {'I','II','III','IV'}: return 'UKG to IV'
    if raw in {'V','VI','VII','VIII','IX'}: return 'V to IX'
    if raw=='X': return 'X'
    if raw in {'XI','XII'}: return 'XI & XII'
    return None

def ensure_fee_structure_seed(session_name='2026-27'):
    changed=False
    for group, terms in DEFAULT_FEE_STRUCTURE_2026_27.items():
        row=FeeStructure.query.filter_by(academic_session=session_name,class_group=group).first()
        if not row:
            row=FeeStructure(academic_session=session_name,class_group=group,term_i=terms[0],term_ii=terms[1],term_iii=terms[2],term_iv=terms[3],total=sum(terms),updated_by='system')
            db.session.add(row); changed=True
    if changed: db.session.commit()

def ensure_fee_payment_windows(session_name='2026-27'):
    defaults = DEFAULT_FEE_PAYMENT_WINDOWS_2026_27 if session_name == '2026-27' else {}
    changed=False
    for key, (label, start, end, fine_from) in defaults.items():
        row=FeePaymentWindow.query.filter_by(academic_session=session_name, term_key=key).first()
        if not row:
            row=FeePaymentWindow(academic_session=session_name, term_key=key, term_label=label, payment_start=start, payment_end=end, fine_from=fine_from, updated_by='system')
            db.session.add(row); changed=True
    if changed:
        db.session.commit()


def fee_structure_for_student(student, session_name='2026-27'):
    group=fee_group_for_student(student)
    return FeeStructure.query.filter_by(academic_session=session_name,class_group=group).first() if group else None

def fee_term_amount(row, term):
    return float(getattr(row, term, 0) or 0) if row else 0.0


# Compatibility / management routes required by the existing V11 templates.
@app.route('/admin/students/<int:sid>/edit', methods=['GET','POST'])
@admin_required
def edit_student(sid):
    student=db.session.get(Student,sid) or abort(404)
    account=student.account
    parents=Parent.query.filter(Parent.account_id.isnot(None)).order_by(Parent.name).all()
    linked_ids={x.parent_id for x in ParentStudent.query.filter_by(student_id=sid).all()}
    if request.method=='POST':
        name=normalize_school_name(request.form.get('name'))
        roll=' '.join((request.form.get('roll_number') or '').strip().upper().split())
        cls=normalize_class(request.form.get('class_name'))
        sec=normalize_section(request.form.get('section'))
        username=(request.form.get('username') or '').strip()
        temp_password=request.form.get('temp_password') or ''
        active=request.form.get('active')=='1'
        if class_number(cls) in {11,12}: sec=''
        if not name or not cls or not roll or not username or (class_number(cls) not in {11,12} and not sec):
            return render_template('student_edit.html',student=student,account=account,parents=parents,linked_parents=linked_ids,class_options=class_list(),error='Name, roll, class, username and the required section are required.'),400
        duplicate=Student.query.filter(Student.id!=sid,Student.class_name==cls,Student.section==sec,Student.roll_number==roll,Student.active.is_(True)).first()
        if duplicate:
            return render_template('student_edit.html',student=student,account=account,parents=parents,linked_parents=linked_ids,class_options=class_list(),error='That roll number already exists in the same class/section.'),409
        other=Account.query.filter(func.lower(Account.username)==username.lower()).first()
        if other and (not account or other.id!=account.id):
            return render_template('student_edit.html',student=student,account=account,parents=parents,linked_parents=linked_ids,class_options=class_list(),error='Username already exists.'),409
        student.name=name; student.roll_number=roll; student.class_name=cls; student.section=sec; student.active=active
        if account:
            account.username=username; account.display_name=name; account.active=active
            if temp_password:
                if len(temp_password)<8: return render_template('student_edit.html',student=student,account=account,parents=parents,linked_parents=linked_ids,class_options=class_list(),error='Temporary password must be at least 8 characters.'),400
                account.password_hash=generate_password_hash(temp_password); account.must_change_password=True
        else:
            password=temp_password or os.getenv('LEGACY_STUDENT_DEFAULT_PASSWORD','ChangeThisImmediately123!')
            account=Account(username=username,password_hash=generate_password_hash(password),role='student',display_name=name,must_change_password=True,active=active); db.session.add(account); db.session.flush(); student.account_id=account.id
        db.session.query(ParentStudent).filter_by(student_id=sid).delete(synchronize_session=False)
        for raw_pid in request.form.getlist('parent_ids'):
            try: pid=int(raw_pid)
            except ValueError: continue
            if db.session.get(Parent,pid): db.session.add(ParentStudent(parent_id=pid,student_id=sid))
        log_audit('student_updated','Student',sid,{'class':cls,'section':sec,'active':active}); db.session.commit()
        flash('Student updated successfully.','success'); return redirect(url_for('admin_students'))
    return render_template('student_edit.html',student=student,account=account,parents=parents,linked_parents=linked_ids,class_options=class_list())

@app.post('/admin/students/<int:sid>/delete')
@admin_required
def delete_student(sid):
    student=db.session.get(Student,sid) or abort(404); student.active=False
    acct=student.account
    if acct:
        student.account_id=None
        db.session.delete(acct)
    log_audit('student_deactivated','Student',sid,{'admission_number':student.admission_number,'account_deleted':bool(acct)})
    db.session.commit()
    flash('Student account deleted; student records and history were preserved. The username can be reused.','success'); return redirect(url_for('admin_students'))

@app.route('/admin/teachers/<int:tid>/edit', methods=['GET','POST'])
@admin_required
def edit_teacher(tid):
    teacher=db.session.get(Teacher,tid) or abort(404); account=teacher.account
    if request.method=='POST':
        name=normalize_school_name(request.form.get('name')); username=(request.form.get('username') or '').strip(); phone=request.form.get('phone','').strip(); email=request.form.get('email','').strip(); active=request.form.get('active')=='1'
        other=Account.query.filter(func.lower(Account.username)==username.lower(), Account.id != (account.id if account else -1)).first()
        if other and not other.active:
            release_inactive_account_username(username); other=None
        if not name or not username or other:
            return render_template('person_edit.html',kind='Teacher',person=teacher,account=account,error='Valid name and unique username are required.'),400
        if not account:
            account=Account(username=username,password_hash=generate_password_hash(os.getenv('LEGACY_STUDENT_DEFAULT_PASSWORD','ChangeThisImmediately123!')),role='teacher',display_name=name,must_change_password=True,active=active); db.session.add(account); db.session.flush(); teacher.account_id=account.id
        else:
            account.username=username; account.display_name=name; account.active=active
        teacher.name=name; teacher.phone=phone; teacher.email=email; teacher.active=active
        log_audit('teacher_updated','Teacher',tid,{'active':active}); db.session.commit(); flash('Teacher updated successfully.','success')
        return redirect(url_for('admin_teachers'))
    return render_template('person_edit.html',kind='Teacher',person=teacher,account=account)

@app.post('/admin/teachers/<int:tid>/delete')
@admin_required
def delete_teacher(tid):
    teacher=db.session.get(Teacher,tid) or abort(404); teacher.active=False
    acct=teacher.account
    if acct:
        teacher.account_id=None
        db.session.delete(acct)
    log_audit('teacher_deactivated','Teacher',tid,{'account_deleted':bool(acct)}); db.session.commit(); flash('Teacher account deleted; teacher records were preserved. The username can be reused.','success')
    return redirect(url_for('admin_teachers'))

@app.route('/admin/parents/<int:pid>/edit', methods=['GET','POST'])
@admin_required
def edit_parent(pid):
    parent=db.session.get(Parent,pid) or abort(404); account=parent.account
    if request.method=='POST':
        name=normalize_school_name(request.form.get('name')); username=(request.form.get('username') or '').strip(); phone=request.form.get('phone','').strip(); email=request.form.get('email','').strip(); active=request.form.get('active')=='1'
        admissions=[x.strip().upper() for x in (request.form.get('child_admissions') or '').replace('\n',',').split(',') if x.strip()]
        students=Student.query.filter(Student.admission_number.in_(admissions),Student.active.is_(True)).all()
        other=Account.query.filter(func.lower(Account.username)==username.lower(), Account.id != (account.id if account else -1)).first()
        if other and not other.active:
            release_inactive_account_username(username); other=None
        if not name or not username or other or not students:
            return render_template('parent_edit.html',parent=parent,account=account,linked_admissions=[s.admission_number for s in students],error='Name, unique username and at least one active child admission number are required.'),400
        if not account:
            account=Account(username=username,password_hash=generate_password_hash(os.getenv('LEGACY_STUDENT_DEFAULT_PASSWORD','ChangeThisImmediately123!')),role='parent',display_name=name,must_change_password=True,active=active); db.session.add(account); db.session.flush(); parent.account_id=account.id
        else:
            account.username=username; account.display_name=name; account.active=active
        parent.name=name; parent.phone=phone; parent.email=email
        db.session.query(ParentStudent).filter_by(parent_id=pid).delete(synchronize_session=False)
        for st in students: db.session.add(ParentStudent(parent_id=pid,student_id=st.id))
        log_audit('parent_updated','Parent',pid,{'children':[s.id for s in students],'active':active}); db.session.commit(); flash('Parent updated successfully.','success')
        return redirect(url_for('admin_accounts',role='parent'))
    linked=[s.admission_number for s in Student.query.join(ParentStudent,ParentStudent.student_id==Student.id).filter(ParentStudent.parent_id==pid).all()]
    return render_template('parent_edit.html',parent=parent,account=account,linked_admissions=linked)

@app.post('/admin/parents/<int:pid>/delete')
@admin_required
def delete_parent(pid):
    parent=db.session.get(Parent,pid) or abort(404)
    acct=parent.account
    if acct:
        parent.account_id=None
        db.session.delete(acct)
    log_audit('parent_deactivated','Parent',pid,{'account_deleted':bool(acct)}); db.session.commit(); flash('Parent account deleted; parent profile and child links were preserved. The username can be reused.','success')
    return redirect(url_for('admin_accounts',role='parent'))

@app.route('/admin/assignments/<int:aid>/edit', methods=['GET','POST'])
@admin_required
def edit_assignment(aid):
    assignment=db.session.get(TeacherAssignment,aid) or abort(404); teachers=Teacher.query.filter_by(active=True).order_by(Teacher.name).all()
    if request.method=='POST':
        try: tid=int(request.form.get('teacher_id'))
        except (TypeError,ValueError): return render_template('assignment_edit.html',assignment=assignment,teachers=teachers,class_options=class_list(),error='Select a valid teacher.'),400
        cls=normalize_class(request.form.get('class_name')); sec=normalize_section(request.form.get('section'))
        if class_number(cls) in {11,12}: sec=''
        teacher=db.session.get(Teacher,tid); conflict=TeacherAssignment.query.filter(TeacherAssignment.class_name==cls,TeacherAssignment.section==sec,TeacherAssignment.id!=aid).first()
        if not teacher or not cls or conflict: return render_template('assignment_edit.html',assignment=assignment,teachers=teachers,class_options=class_list(),error='That class/section already has a class teacher, or the teacher is invalid.'),409
        assignment.teacher_id=tid; assignment.class_name=cls; assignment.section=sec; log_audit('teacher_assignment_updated','TeacherAssignment',aid,{'teacher_id':tid,'class_name':cls,'section':sec}); db.session.commit(); flash('Class-teacher assignment updated.','success'); return redirect(url_for('assignments'))
    return render_template('assignment_edit.html',assignment=assignment,teachers=teachers,class_options=class_list())

@app.post('/admin/assignments/<int:aid>/delete')
@admin_required
def delete_assignment(aid):
    assignment=db.session.get(TeacherAssignment,aid) or abort(404); db.session.delete(assignment); log_audit('teacher_assignment_deleted','TeacherAssignment',aid); db.session.commit(); flash('Assignment removed.','success'); return redirect(url_for('assignments'))

@app.post('/admin/results/publish')
@admin_required
def publish_results():
    try: exam_id=int(request.form.get('exam_id'))
    except (TypeError,ValueError): return redirect(url_for('results'))
    exam=db.session.get(Exam,exam_id) or abort(404); cls=normalize_class(request.form.get('class_name')); sec=normalize_section(request.form.get('section'))
    if not exam.is_final:
        flash('Only the final examination can be published to students and parents.','error'); return redirect(url_for('results'))
    if not cls: flash('A valid class is required.','error'); return redirect(url_for('results'))
    row=ResultPublication.query.filter_by(exam_id=exam.id,class_name=cls,section=sec).first() or ResultPublication(exam_id=exam.id,class_name=cls,section=sec)
    db.session.add(row); row.published=True; row.published_at=datetime.utcnow(); row.published_by=current_account().username; log_audit('results_published','ResultPublication',row.id,{'exam':exam.name,'class_name':cls,'section':sec}); db.session.commit(); flash(f'{exam.name} results published for {cls}{("-"+sec) if sec else ""}.','success'); return redirect(url_for('results'))

@app.post('/admin/results/unpublish')
@admin_required
def unpublish_results():
    try: exam_id=int(request.form.get('exam_id'))
    except (TypeError,ValueError): return redirect(url_for('results'))
    exam=db.session.get(Exam,exam_id) or abort(404); cls=normalize_class(request.form.get('class_name')); sec=normalize_section(request.form.get('section'))
    if not exam.is_final:
        flash('Only the final examination publication is managed here.','error'); return redirect(url_for('results'))
    row=ResultPublication.query.filter_by(exam_id=exam.id,class_name=cls,section=sec).first()
    if row:
        row.published=False; row.published_at=None; row.published_by=current_account().username; log_audit('results_unpublished','ResultPublication',row.id,{'exam':exam.name,'class_name':cls,'section':sec}); db.session.commit()
    flash(f'{exam.name} results unpublished for {cls}{("-"+sec) if sec else ""}.','info'); return redirect(url_for('results'))

@app.route('/admin/test-clock', methods=['GET','POST'])
@admin_required
def test_clock():
    clock=db.session.get(SchoolClock,1)
    if clock is None: clock=SchoolClock(id=1); db.session.add(clock); db.session.flush()
    if request.method=='POST':
        if request.form.get('action')=='clear':
            clock.override_time=None; log_audit('school_clock_cleared','SchoolClock',1); db.session.commit(); flash('School clock restored to real Asia/Kolkata time.','success')
        else:
            raw=(request.form.get('override_time') or '').strip()
            try: parsed=datetime.strptime(raw,'%H:%M').time()
            except ValueError: return render_template('test_clock.html',clock=clock,school_time=school_time(),status=attendance_status_for_time(),error='Use a valid 24-hour time such as 08:45.'),400
            clock.override_time=parsed.strftime('%H:%M'); log_audit('school_clock_set','SchoolClock',1,{'override_time':clock.override_time}); db.session.commit(); flash(f'Test school time set to {clock.override_time}.','success')
        return redirect(url_for('test_clock'))
    return render_template('test_clock.html',clock=clock,school_time=school_time(),status=attendance_status_for_time())

# Fees are deliberately manual: the system never creates invoices automatically.
def current_academic_session():
    setting=SchoolSetting.query.filter_by(key='academic_session').first()
    return (setting.value.strip() if setting and setting.value else os.getenv('ACADEMIC_SESSION','2026-27'))

def fee_status_for_student(student, session_name=None):
    session_name=session_name or current_academic_session(); row=fee_structure_for_student(student,session_name); result=[]
    for term_key,label in FEE_TERM_LABELS:
        amount=fee_term_amount(row,term_key)
        if amount<=0: continue
        title=f'School Fee {session_name} • {label}'
        invs=FeeInvoice.query.filter_by(student_id=student.id).filter(FeeInvoice.title==title).order_by(FeeInvoice.created_at.desc()).all()
        paid=sum(sum(float(p.amount or 0) for p in (inv.payments or [])) for inv in invs)
        result.append({'term_key':term_key,'term_label':label,'amount':amount,'paid':paid,'balance':max(0,amount-paid),'state':'paid' if paid>=amount else ('partial' if paid>0 else 'unpaid'),'invoice_id':invs[0].id if invs else None})
    return result

def maybe_create_fee_reminders(student_ids=None, session_name=None):
    # Dashboard pages can call this helper more than once; never duplicate work
    # inside the same request.
    cache_key=('fee_reminders', tuple(sorted(int(x) for x in (student_ids or []))), session_name or '')
    seen=getattr(g,'fee_reminder_checks',set())
    if cache_key in seen:
        return 0
    seen.add(cache_key); g.fee_reminder_checks=seen
    session_name=session_name or current_academic_session(); ids=list(student_ids or [])
    students=Student.query.filter(Student.active.is_(True),Student.id.in_(ids)).all() if ids else Student.query.filter(Student.active.is_(True)).all(); windows={w.term_key:w for w in FeePaymentWindow.query.filter_by(academic_session=session_name).all()}; today=school_date(); created=0
    for st in students:
        if not ParentStudent.query.filter_by(student_id=st.id).first(): continue
        for item in fee_status_for_student(st,session_name):
            w=windows.get(item['term_key'])
            if not w or today < w.fine_from or item['state']=='paid': continue
            for link in ParentStudent.query.filter_by(student_id=st.id).all():
                title=f'Fee reminder • {st.admission_number} • {item["term_key"]}'
                if Announcement.query.filter_by(parent_id=link.parent_id,title=title).first(): continue
                db.session.add(Announcement(title=title,message=f'Fee reminder for {st.name}: {item["term_label"]} payment of ₹{item["amount"]:,.0f} is still {item["state"]}. Please contact the school office for payment verification.',audience='parents',published=True,created_by='system',published_at=datetime.utcnow(),parent_id=link.parent_id)); created+=1
    if created: db.session.commit()
    return created

def fee_lookup_candidates(args):
    q=Student.query.filter(Student.active.is_(True)); exact=(args.get('admin_no') or args.get('admission') or '').strip().upper(); name=(args.get('name') or '').strip(); cls=normalize_class(args.get('class_name')); sec=normalize_section(args.get('section')); roll=(args.get('roll') or '').strip().upper()
    if exact: q=q.filter(Student.admission_number==exact)
    if name: q=q.filter(Student.name.ilike(f'%{name}%'))
    if cls: q=q.filter(Student.class_name==cls)
    if sec: q=q.filter(Student.section==sec)
    if roll: q=q.filter(Student.roll_number==roll)
    return student_order(q).limit(100).all()

def render_fee_structure_page(session_name):
    rows=FeeStructure.query.filter_by(academic_session=session_name).order_by(FeeStructure.id).all(); windows=FeePaymentWindow.query.filter_by(academic_session=session_name).order_by(FeePaymentWindow.payment_start).all(); docs=FeeStructureDocument.query.filter_by(academic_session=session_name).order_by(FeeStructureDocument.uploaded_at.desc()).all()
    parts=["<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>Fee Structure</title></head><body style='font-family:Arial,sans-serif;max-width:1000px;margin:40px auto;padding:20px'>",'<h1>DAV PS KKP - Fee Structure</h1>',f'<p>Academic session: <strong>{session_name}</strong></p>',"<table border='1' cellpadding='8' cellspacing='0' width='100%'><tr><th>Class</th><th>Term I</th><th>Term II</th><th>Term III</th><th>Term IV</th><th>Total</th></tr>"]
    for r in rows: parts.append(f'<tr><td>{r.class_group}</td><td>₹{r.term_i:,.0f}</td><td>₹{r.term_ii:,.0f}</td><td>₹{r.term_iii:,.0f}</td><td>₹{r.term_iv:,.0f}</td><td>₹{r.total:,.0f}</td></tr>')
    parts.append("</table><h2>Payment windows</h2><table border='1' cellpadding='8' cellspacing='0' width='100%'><tr><th>Term</th><th>Without fine</th><th>Fine from</th></tr>")
    for w in windows: parts.append(f'<tr><td>{w.term_label}</td><td>{w.payment_start.strftime("%d.%m.%Y")} - {w.payment_end.strftime("%d.%m.%Y")}</td><td>{w.fine_from.strftime("%d.%m.%Y")}</td></tr>')
    parts.append(f"</table><h2>Upload annual fee document</h2><form method='post' enctype='multipart/form-data'><input type='file' name='fee_document' accept='.pdf,.png,.jpg,.jpeg,.webp' required><button type='submit'>Upload</button></form><p>Stored documents: {len(docs)}</p><p><a href='/admin/fees/ledger'>Fee ledger</a></p></body></html>")
    return ''.join(parts)

@app.route('/admin/fees/structure', methods=['GET','POST'])
@admin_required
def admin_fee_structure():
    session_name=(request.args.get('session') or current_academic_session()).strip()[:30]
    if session_name=='2026-27': ensure_fee_structure_seed(session_name); ensure_fee_payment_windows(session_name)
    if request.method=='POST':
        upload=request.files.get('fee_document')
        if not upload or not upload.filename: return 'Fee document is required.',400
        raw=upload.read(); mimetype=(upload.mimetype or '').lower()
        if len(raw)>10*1024*1024: return 'Fee document must be 10 MB or smaller.',400
        if mimetype not in {'application/pdf','image/png','image/jpeg','image/webp'}: return 'Upload a PDF, PNG, JPG, or WEBP document.',400
        db.session.add(FeeStructureDocument(academic_session=session_name,filename=Path(upload.filename).name[:255],mimetype=mimetype,file_data=raw,uploaded_by=current_account().username)); log_audit('fee_document_uploaded','FeeStructureDocument',extra={'session':session_name,'filename':upload.filename}); db.session.commit(); return redirect(url_for('admin_fee_structure',session=session_name))
    return render_fee_structure_page(session_name)

@app.get('/admin/fees/ledger')
@admin_required
def admin_fee_ledger():
    session_name=(request.args.get('session') or current_academic_session()).strip()[:30]; candidates=fee_lookup_candidates(request.args); selected=candidates[0] if len(candidates)==1 else None
    if selected: maybe_create_fee_reminders([selected.id],session_name)
    parts=["<!doctype html><html><head><meta charset='utf-8'><title>Fee Ledger</title></head><body style='font-family:Arial,sans-serif;max-width:1100px;margin:40px auto;padding:20px'><h1>Fee Ledger</h1><form method='get'><input name='admin_no' placeholder='Admission / admin no.' value='",str(request.args.get('admin_no','')),"'><input name='name' placeholder='Name' value='",str(request.args.get('name','')),"'><input name='class_name' placeholder='Class'><input name='section' placeholder='Section'><input name='roll' placeholder='Roll'><button>Search</button></form>"]
    if not selected:
        parts.append('<h2>Matches</h2><ul>'+''.join(f'<li><a href="?admin_no={s.admission_number}">{s.name} - {s.admission_number} - {s.class_name}{("-"+s.section) if s.section else ""} - Roll {s.roll_number or "-"}</a></li>' for s in candidates)+'</ul>')
    else:
        parts.append(f'<h2>{selected.name}</h2><p>Admission: {selected.admission_number} - {selected.class_name}{("-"+selected.section) if selected.section else ""} - Roll: {selected.roll_number or "-"}</p><table border="1" cellpadding="8" cellspacing="0"><tr><th>Term</th><th>Amount</th><th>Paid</th><th>Balance</th><th>Status</th></tr>')
        for x in fee_status_for_student(selected,session_name): parts.append(f'<tr><td>{x["term_label"]}</td><td>₹{x["amount"]:,.0f}</td><td>₹{x["paid"]:,.0f}</td><td>₹{x["balance"]:,.0f}</td><td>{x["state"]}</td></tr>')
        parts.append('</table>')
    parts.append("<p><a href='/admin/fees/structure'>Fee structure</a> - <a href='/dashboard'>Dashboard</a></p></body></html>"); return ''.join(parts)

@app.get('/admin/fees/document/<int:doc_id>')
@admin_required
def fee_document(doc_id):
    doc=db.session.get(FeeStructureDocument,doc_id) or abort(404); return send_file(BytesIO(doc.file_data),download_name=doc.filename,mimetype=doc.mimetype)

@app.post('/admin/fees/payment')
@admin_required
def record_fee_payment():
    try: invoice_id=int(request.form.get('invoice_id')); amount=float(request.form.get('amount')); method=(request.form.get('method') or 'offline').strip()[:40]; receipt=(request.form.get('receipt_no') or '').strip()[:80]
    except (TypeError,ValueError): return jsonify({'error':'Invoice, amount and receipt are required.'}),400
    inv=db.session.get(FeeInvoice,invoice_id) or abort(404)
    if amount<=0 or not receipt: return jsonify({'error':'Payment amount and receipt number are required.'}),400
    already=sum(float(p.amount or 0) for p in (inv.payments or [])); remaining=max(0,float(inv.amount_due or 0)-already)
    if amount>remaining: return jsonify({'error':f'Payment exceeds remaining balance ₹{remaining:,.2f}.'}),400
    if FeePayment.query.filter_by(receipt_no=receipt).first(): return jsonify({'error':'Receipt number already exists.'}),409
    pay=FeePayment(invoice_id=inv.id,amount=amount,method=method or 'offline',receipt_no=receipt,received_by=current_account().username); db.session.add(pay); db.session.flush(); total=already+amount; inv.status='paid' if total>=float(inv.amount_due or 0) else 'partial'; log_audit('fee_payment_verified','FeePayment',pay.id,{'invoice_id':invoice_id,'amount':amount,'receipt_no':receipt}); db.session.commit(); return jsonify({'ok':True,'status':inv.status,'paid':total,'balance':max(0,float(inv.amount_due or 0)-total)})




def extract_calendar_upload_text(upload):
    if upload is None: return ''
    raw=upload.read(); name=(upload.filename or '').lower()
    if name.endswith('.pdf'):
        if PdfReader is None: raise ValueError('PDF support is not installed on this server.')
        return '\n'.join((page.extract_text() or '') for page in PdfReader(BytesIO(raw)).pages)
    if pytesseract is None or cv2 is None or np is None: raise ValueError('Image calendar OCR is unavailable; paste calendar text instead.')
    arr=np.frombuffer(raw,dtype=np.uint8); img=cv2.imdecode(arr,cv2.IMREAD_COLOR)
    if img is None: raise ValueError('The calendar image could not be decoded.')
    return pytesseract.image_to_string(img)

def parse_calendar_text(text_value):
    pattern=re.compile(r'(?<!\d)(\d{1,2})[./-](\d{1,2})[./-](\d{4})(?!\d)'); rows={}; value=str(text_value or '')
    for m in pattern.finditer(value):
        try: day=date(int(m.group(3)),int(m.group(2)),int(m.group(1)))
        except ValueError: continue
        context=value[max(0,m.start()-80):min(len(value),m.end()+120)].lower(); off=any(k in context for k in ('holiday','closed','vacation','non-working','non working')); rows[day.isoformat()]={'date':day.isoformat(),'is_working':not off,'reason':'Holiday' if off else 'Imported calendar date'}
    return sorted(rows.values(),key=lambda x:x['date'])

@app.get('/admin/calendar')
@admin_required
def admin_calendar_page():
    return render_template('calendar.html',calendar_readonly=False)

@app.get('/teacher/calendar')
@staff_required
def teacher_calendar_page():
    if current_account().role != 'teacher': abort(403)
    return render_template('calendar.html',calendar_readonly=True)

@app.get('/calendar')
@login_required()
def calendar_view_redirect():
    if current_account().role=='admin': return redirect(url_for('admin_calendar_page'))
    if current_account().role=='teacher': return redirect(url_for('teacher_calendar_page'))
    return render_template('calendar.html',calendar_readonly=True)

@app.get('/api/calendar')
@login_required()
def calendar_api():
    year=request.args.get('year',school_date().year,type=int); month=request.args.get('month',school_date().month,type=int)
    if not 1<=month<=12 or not 1900<=year<=2200: return jsonify({'error':'Invalid calendar month'}),400
    start=datetime(year,month,1).date(); nxt=datetime(year+1,1,1).date() if month==12 else datetime(year,month+1,1).date(); end=nxt-timedelta(days=1)
    overrides={r.date:r for r in SchoolCalendar.query.filter(SchoolCalendar.date>=start,SchoolCalendar.date<=end).all()}
    rows=[]; day=start
    while day<=end:
        ov=overrides.get(day); rows.append({'date':day.isoformat(),'is_working':bool(ov.is_working) if ov else weekly_default_is_working(day),'reason':ov.reason if ov else '','override':bool(ov)}); day+=timedelta(days=1)
    return jsonify(rows)

@app.post('/api/calendar')
@admin_required
def calendar_set():
    data=request.get_json(silent=True) or {}; raw=str(data.get('date') or '')
    try: day=datetime.strptime(raw,'%Y-%m-%d').date()
    except ValueError: return jsonify({'error':'Valid date is required'}),400
    row=SchoolCalendar.query.filter_by(date=day).first()
    if row is None:
        row=SchoolCalendar(date=day)
        db.session.add(row)
    row.is_working=parse_bool(data.get('is_working'), True)
    row.reason=(data.get('reason') or '').strip()[:255] or None
    now_utc=datetime.utcnow()
    if not row.created_at: row.created_at=now_utc
    row.updated_at=now_utc
    try:
        db.session.flush(); log_audit('calendar_update','SchoolCalendar',row.id,{'date':day.isoformat(),'is_working':row.is_working,'reason':row.reason}); db.session.commit()
    except IntegrityError:
        db.session.rollback(); return jsonify({'error':'Calendar update conflicted with another request. Refresh and try again.'}),409
    return jsonify({'message':'Calendar updated','date':day.isoformat(),'is_working':row.is_working,'reason':row.reason or ''})

@app.post('/api/calendar/reset')
@admin_required
def calendar_reset():
    data=request.get_json(silent=True) or {}; raw=str(data.get('date') or '')
    try: day=datetime.strptime(raw,'%Y-%m-%d').date()
    except ValueError: return jsonify({'error':'Valid date is required'}),400
    row=SchoolCalendar.query.filter_by(date=day).first()
    if row:
        rid=row.id; db.session.delete(row); log_audit('calendar_reset','SchoolCalendar',rid,{'date':day.isoformat()}); db.session.commit()
    return jsonify({'date':day.isoformat(),'is_working':weekly_default_is_working(day)})

@app.post('/api/calendar/bulk')
@admin_required
def calendar_bulk():
    data=request.get_json(silent=True) or {}; dates=data.get('dates') or []; is_working=parse_bool(data.get('is_working'), True); reason=(data.get('reason') or '').strip()[:255] or None; changed=0
    try:
        for raw in dates:
            day=datetime.strptime(str(raw),'%Y-%m-%d').date(); row=SchoolCalendar.query.filter_by(date=day).first() or SchoolCalendar(date=day); row.is_working=is_working; row.reason=reason; row.updated_at=datetime.utcnow(); db.session.add(row); changed+=1
        log_audit('calendar_bulk_update','SchoolCalendar',extra={'changed':changed,'is_working':is_working,'reason':reason}); db.session.commit()
    except (TypeError,ValueError): db.session.rollback(); return jsonify({'error':'Every date must use YYYY-MM-DD'}),400
    return jsonify({'message':f'Updated {changed} calendar date(s)'})

@app.post('/api/calendar/import')
@admin_required
def calendar_import():
    upload=request.files.get('calendar_file'); uploaded_text=(request.form.get('calendar_text') or '').strip()
    if not upload and not uploaded_text: return jsonify({'error':'Choose a school calendar PDF or image first.'}),400
    try:
        text_value=uploaded_text or extract_calendar_upload_text(upload); rows=parse_calendar_text(text_value)
        if not rows: return jsonify({'error':'No dates could be detected. You can also edit days manually.'}),400
        return jsonify({'dates':rows,'working_count':sum(1 for r in rows if r['is_working']),'non_working_count':sum(1 for r in rows if not r['is_working']),'text_preview':text_value[:2000]})
    except ValueError as exc: return jsonify({'error':str(exc)}),400
    except Exception as exc:
        app.logger.exception('Calendar import failed'); return jsonify({'error':f'Calendar import failed: {exc}'}),500

@app.post('/api/calendar/import/apply')
@admin_required
def calendar_import_apply():
    rows=(request.get_json(silent=True) or {}).get('dates') or []; changed=0
    try:
        for item in rows:
            day=datetime.strptime(str(item.get('date') or ''),'%Y-%m-%d').date(); row=SchoolCalendar.query.filter_by(date=day).first() or SchoolCalendar(date=day); row.is_working=parse_bool(item.get('is_working'), True); row.reason=(item.get('reason') or '').strip()[:255] or None; row.updated_at=datetime.utcnow(); db.session.add(row); changed+=1
        log_audit('calendar_import','SchoolCalendar',extra={'changed':changed}); db.session.commit()
    except (TypeError,ValueError): db.session.rollback(); return jsonify({'error':'One or more detected calendar dates are invalid.'}),400
    return jsonify({'message':f'Applied {changed} calendar date(s)'})

@app.get('/audit')
@audit_network_required()
def audit_login():
    if session.get('audit_auth'): return redirect(url_for('audit_view'))
    return render_template('audit_login.html')

@app.post('/audit')
@audit_network_required()
def audit_auth():
    if secrets.compare_digest(request.form.get('username',''),os.getenv('AUDIT_USERNAME','auditor')) and secrets.compare_digest(request.form.get('password',''),os.getenv('AUDIT_PASSWORD','ChangeAuditPasswordImmediately!')):
        session['audit_auth']=True; session['audit_username']=request.form.get('username'); return redirect(url_for('audit_view'))
    return render_template('audit_login.html',error='Invalid audit credentials.'),401

@app.get('/audit/view')
@audit_network_required()
def audit_view():
    if not session.get('audit_auth'): return redirect(url_for('audit_login'))
    return render_template('audit_view.html',events=AuditEvent.query.order_by(AuditEvent.id.desc()).limit(500).all())

def academic_percentage_map(student_ids):
    ids=[int(x) for x in student_ids if x is not None]
    if not ids: return {}
    rows=(db.session.query(Mark.student_id, func.coalesce(func.sum(Mark.marks),0), func.coalesce(func.sum(Mark.max_marks),0))
          .filter(Mark.student_id.in_(ids), Mark.marks.isnot(None))
          .group_by(Mark.student_id).all())
    return {sid: round((float(got)/float(mx))*100,2) if mx else 0 for sid,got,mx in rows}

@app.post('/api/ai')
@login_required()
def ai_api():
    acct=current_account(); data=request.get_json(silent=True) or {}; question=str(data.get('question') or '').strip()
    if not question: return jsonify({'error':'Question is required.'}),400
    students=student_order(allowed_students_for_account(acct)).limit(5000).all(); start,end=school_year_bounds(); today=min(school_date(),end); ids=[s.id for s in students]
    return jsonify({'ok':True,'answer':_local_ai_fallback(question,acct,students,attendance_percentage_map(ids,start,today),academic_percentage_map(ids)),'source':'local-school-data','scope':acct.role,'student_count':len(students)})


def initialize_application():
    with app.app_context():
        uri=app.config['SQLALCHEMY_DATABASE_URI']; auto_create=os.getenv('AUTO_CREATE_SCHEMA','').lower() in {'1','true','yes','on'} or uri.startswith('sqlite:///'); auto_seed=os.getenv('AUTO_SEED','').lower() in {'1','true','yes','on'}
        if auto_create: db.create_all()
        if uri.startswith('sqlite:///'):
            try:
                db.session.execute(text('PRAGMA journal_mode=WAL'))
                db.session.execute(text('PRAGMA synchronous=NORMAL'))
                db.session.execute(text('PRAGMA foreign_keys=ON'))
                db.session.commit()
            except Exception:
                db.session.rollback()
        try: ensure_announcement_schema()
        except Exception:
            if not uri.startswith('sqlite:///'): raise
        if auto_seed: seed_data()
        try:
            if current_academic_session()=='2026-27': ensure_fee_structure_seed('2026-27'); ensure_fee_payment_windows('2026-27')
        except Exception:
            db.session.rollback(); app.logger.exception('Optional fee bootstrap failed')

initialize_application()


if __name__ == '__main__':
    port = int(os.getenv('PORT', '5000'))
    ssl_context=None
    use_https=app.config.get('LOCAL_HTTPS',False)
    cert_file=app.config.get('SSL_CERT_FILE') or ''
    key_file=app.config.get('SSL_KEY_FILE') or ''
    if use_https:
        if cert_file and key_file:
            ssl_context=(cert_file,key_file)
        else:
            ssl_context='adhoc'
        app.logger.warning('Local HTTPS enabled. Browser/phone may warn about the development certificate.')
    app.run(host='0.0.0.0', port=port, debug=False, ssl_context=ssl_context)
