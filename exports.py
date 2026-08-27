from io import BytesIO
from pathlib import Path
import json, base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import mm


def build_xlsx(rows, title='School Report'):
    wb = Workbook(); ws = wb.active; ws.title = 'Report'
    ws.append([title]); ws.append([])
    if rows:
        headers = list(rows[0].keys()); ws.append(headers)
        for r in rows: ws.append([r.get(h) for h in headers])
        for c in ws[3]: c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='EAF0FF')
        for col in ws.columns:
            width = max(len(str(x.value or '')) for x in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(width, 35)
    out = BytesIO(); wb.save(out); out.seek(0); return out


def build_pdf(rows, title='School Report', subtitle=''):
    out = BytesIO(); doc = SimpleDocTemplate(out, pagesize=(A4[1], A4[0]), rightMargin=18,leftMargin=18,topMargin=18,bottomMargin=18)
    styles = getSampleStyleSheet(); story=[Paragraph(title, styles['Title'])]
    if subtitle: story.append(Paragraph(subtitle, styles['Normal']))
    story.append(Spacer(1,8))
    if rows:
        headers=list(rows[0].keys()); data=[headers]+[[str(r.get(h,'')) for h in headers] for r in rows]
        t=Table(data, repeatRows=1)
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#EAF0FF')),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#D7DCE6')),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('FONTSIZE',(0,0),(-1,-1),8)]))
        story.append(t)
    doc.build(story); out.seek(0); return out


def _grade(pct):
    if pct >= 91: return 'A1'
    if pct >= 81: return 'A2'
    if pct >= 71: return 'B1'
    if pct >= 61: return 'B2'
    if pct >= 51: return 'C1'
    if pct >= 41: return 'C2'
    if pct >= 33: return 'D'
    return 'E'


def _decode_signature(data, width=34*mm, height=14*mm):
    if not data or not str(data).startswith('data:image'):
        return None
    try:
        raw = base64.b64decode(str(data).split(',',1)[1])
        return Image(BytesIO(raw), width=width, height=height, preserveAspectRatio=True, mask='auto')
    except Exception:
        return None


def _header_story(story, student, session_name, logo_path=None):
    styles = getSampleStyleSheet()
    title = ParagraphStyle('cardtitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=19, leading=21, alignment=TA_CENTER, textColor=colors.HexColor('#173A63'), spaceAfter=2)
    sub = ParagraphStyle('cardsmall', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5, leading=10, alignment=TA_CENTER, textColor=colors.HexColor('#58667A'))
    brand = []
    if logo_path:
        try: brand.append(Image(str(logo_path), width=22*mm, height=22*mm, preserveAspectRatio=True, mask='auto'))
        except Exception: pass
    brand.extend([Paragraph('D.A.V. PUBLIC SCHOOL', title), Paragraph('ACADEMIC PROGRESS REPORT', sub), Paragraph(f'Academic Session: <b>{session_name}</b>', sub)])
    if len(brand) == 4:
        t = Table([[brand[0], brand[1], brand[2], brand[3]]], colWidths=[26*mm, 55*mm, 58*mm, 38*mm])
    else:
        t = Table([[Paragraph('D.A.V. PUBLIC SCHOOL', title), Paragraph('ACADEMIC PROGRESS REPORT', sub), Paragraph(f'Academic Session: <b>{session_name}</b>', sub)]], colWidths=[65*mm, 65*mm, 55*mm])
    t.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER'),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
    story.append(t); story.append(Spacer(1,4))


def _boxed_table(data, widths, header_rows=1, fontsize=7.5, header_bg='#EAF0FF'):
    t=Table(data,colWidths=widths,repeatRows=header_rows)
    st=[('GRID',(0,0),(-1,-1),0.35,colors.HexColor('#C9D2DE')),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('FONTSIZE',(0,0),(-1,-1),fontsize),('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4)]
    if header_rows:
        st += [('BACKGROUND',(0,0),(-1,header_rows-1),colors.HexColor(header_bg)),('FONTNAME',(0,0),(-1,header_rows-1),'Helvetica-Bold')]
    t.setStyle(TableStyle(st)); return t


def build_report_card(student, session_name, attendance, term1, term2, subjects, exams, marks_by_subject, teacher_signature=None, config=None, logo_path=None, assessment_by_subject=None):
    """Build a complete, polished school progress report while preserving V11 data structures."""
    config=config or {}
    assessment_by_subject=assessment_by_subject or {}
    co=config.get('co_scholastic',{}) or {}
    dis=config.get('discipline',{}) or {}
    health=config.get('health',{}) or {}
    layout=config.get('layout',{}) or {}
    report_title=layout.get('report_title') or 'STUDENT PROGRESS REPORT'
    attendance_title=layout.get('attendance_title') or 'ATTENDANCE SUMMARY'
    scholastic_title=layout.get('scholastic_title') or 'SCHOLASTIC PERFORMANCE'
    development_title=layout.get('development_title') or 'CO-SCHOLASTIC & PERSONAL DEVELOPMENT'
    guide_title=layout.get('guide_title') or 'PROGRESS REPORT GUIDE'
    class_details_title=layout.get('class_details_title') or 'STUDENT & SCHOOL DETAILS'
    pass_rule=layout.get('pass_rule') or '33% in every subject.'
    teacher_signature_label=layout.get('teacher_signature_label') or 'Class Teacher Signature'
    principal_signature_label=layout.get('principal_signature_label') or 'Principal Signature'
    parent_signature_label=layout.get('parent_signature_label') or 'Parent Signature'

    out=BytesIO()
    doc=SimpleDocTemplate(out,pagesize=A4,rightMargin=9*mm,leftMargin=9*mm,topMargin=8*mm,bottomMargin=8*mm)
    styles=getSampleStyleSheet()
    body=ParagraphStyle('rb',parent=styles['BodyText'],fontSize=8.0,leading=9.8,textColor=colors.HexColor('#26344A'))
    small=ParagraphStyle('rs',parent=body,fontSize=6.9,leading=8.2)
    tiny=ParagraphStyle('rt',parent=body,fontSize=6.2,leading=7.2)
    h=ParagraphStyle('rh',parent=styles['Heading2'],fontSize=11.3,leading=13,textColor=colors.HexColor('#173A63'),spaceBefore=3,spaceAfter=5)
    h_center=ParagraphStyle('rhc',parent=h,alignment=TA_CENTER)
    section_note=ParagraphStyle('rsn',parent=small,textColor=colors.HexColor('#5A6B80'),spaceAfter=4)

    def safe(v):
        return str(v).strip() if v not in (None,'') else ''
    def esc(v):
        return safe(v).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('\n','<br/>')
    def para(v,sty=body):
        return Paragraph(esc(v),sty) if safe(v) else ''
    def fmt_num(v):
        try:
            f=float(v)
            return str(int(f)) if f.is_integer() else f'{f:g}'
        except Exception:
            return safe(v)

    story=[]
    _header_story(story,student,session_name,logo_path)
    story.append(Paragraph(report_title,h_center))
    story.append(Paragraph('Official academic record • Student copy',section_note))

    # Complete student identity block. These are automatic fields, not extra entry fields.
    class_label=student.class_name + (f' - {student.section}' if safe(student.section) else '')
    identity=[
        ('Student Name',student.name),('Admission No.',student.admission_number),('Roll No.',student.roll_number),('Class / Section',class_label),
        ('Date of Birth',getattr(student,'date_of_birth',None).strftime('%d.%m.%Y') if getattr(student,'date_of_birth',None) else ''),
        ('House',config.get('house','')),('1st Language','English'),
        ('2nd Language',safe(getattr(student,'second_language','')).title()),('3rd Language',safe(getattr(student,'third_language','')).title()),
        ('Mother',getattr(student,'mother_name','')),('Father',getattr(student,'father_name','')),('Previous School',getattr(student,'previous_school','')),
        ('Academic Session',config.get('academic_session') or session_name),('Result Date',config.get('date_result','')),
    ]
    identity=[(a,b) for a,b in identity if safe(b)]
    if identity:
        rows=[]
        for i in range(0,len(identity),2):
            pair=identity[i:i+2]
            row=[]
            for label,val in pair:
                row.extend([Paragraph(label,tiny),Paragraph(esc(val),small)])
            if len(pair)==1: row.extend(['',''])
            rows.append(row)
        story.append(_boxed_table(rows,[30*mm,62*mm,30*mm,62*mm],0,6.7))

    story.append(Spacer(1,5))
    story.append(Paragraph(attendance_title,h))
    story.append(Paragraph('Attendance figures are based on working-school days recorded in the system. Late entries are included in Present for the attendance percentage.',section_note))
    att_rows=[['Period','Working Days','Present','Absent','Attendance %']]
    for label,x in [('Term I',term1 or {}),('Term II',term2 or {}),('Current / Year',attendance or {})]:
        vals=[x.get('working_days',0),x.get('present',0),x.get('absent',0),f"{float(x.get('percentage',0) or 0):.2f}%"]
        att_rows.append([label,*vals])
    story.append(_boxed_table([[Paragraph(str(x),small) for x in row] for row in att_rows],[38*mm,34*mm,31*mm,31*mm,36*mm],1,7.0))

    if safe(config.get('class_teacher_name')):
        story.append(Spacer(1,5))
        story.append(Paragraph(class_details_title,h))
        story.append(_boxed_table([
            [Paragraph('Class Teacher',small),Paragraph(esc(config.get('class_teacher_name')),body)],
            [Paragraph('Next Academic Session',small),Paragraph(esc(config.get('next_academic_session','')),body)],
        ],[45*mm,125*mm],0,7.2))

    # Assessment scheme overview keeps all exam metadata visible.
    story.append(Spacer(1,5))
    story.append(Paragraph('Assessment Scheme',h))
    exam_rows=[['Exam / Assessment','Maximum Marks','Final Exam?']]
    for e in exams:
        exam_rows.append([e.name,e.max_marks,'Yes' if e.is_final else 'No'])
    if len(exam_rows)>1:
        story.append(_boxed_table([[Paragraph(str(x),small) for x in row] for row in exam_rows],[95*mm,35*mm,40*mm],1,7.0))

    story.append(PageBreak())
    _header_story(story,student,session_name,logo_path)
    story.append(Paragraph(scholastic_title,h))
    story.append(Paragraph('Every recorded examination is shown below. Each cell contains obtained marks / maximum marks.',section_note))

    if exams:
        # Keep all exams visible without creating an unreadably wide weighted table.
        exam_cols=[e for e in exams]
        widths=[39*mm]+[min(24, max(15, 140/len(exam_cols)))*mm for _ in exam_cols]
        exam_data=[[Paragraph('Subject',tiny)]+[Paragraph(f'{e.name}<br/><b>/{e.max_marks}</b>',tiny) for e in exam_cols]]
        weighted_totals=[]
        for subj in subjects:
            row=[Paragraph(esc(subj.name),small)]
            total_got=0; total_max=0
            for e in exam_cols:
                m=marks_by_subject.get((subj.code,e.id))
                if m and m.marks is not None:
                    row.append(Paragraph(f'{fmt_num(m.marks)} / {fmt_num(m.max_marks)}',tiny))
                    total_got += float(m.marks); total_max += float(m.max_marks or e.max_marks or 0)
                else:
                    row.append(Paragraph('—',tiny))
            if total_max:
                weighted_totals.append((subj,total_got,total_max))
            exam_data.append(row)
        if len(exam_data)>1:
            story.append(_boxed_table(exam_data,widths,1,6.0))

    story.append(Spacer(1,7))
    story.append(Paragraph('Final Subject-wise Assessment',h))
    story.append(Paragraph('Internal Assessment is based on the best-two PT contribution plus Multiple Assessment, Subject Enrichment and Portfolio, capped at /20; Final Examination is shown /80.',section_note))

    pt=[e for e in exams if str(e.name).upper() in {'PT-1','PT-2','PT-3'}]
    finals=[e for e in exams if e.is_final or str(e.name).lower().startswith('final')]
    final_headers=['Subject','Best 2 PT /5','Internal /20','Final /80','Total /100','Grade']
    final_data=[[Paragraph(x,tiny) for x in final_headers]]
    overall_total=0.0; overall_count=0
    for subj in subjects:
        pt_scores=[]
        for e in pt:
            m=marks_by_subject.get((subj.code,e.id))
            if m and m.marks is not None and m.max_marks:
                pt_scores.append((float(m.marks)/float(m.max_marks))*5.0)
        best2=sum(sorted(pt_scores,reverse=True)[:2])/len(sorted(pt_scores,reverse=True)[:2]) if pt_scores else 0.0
        comp=assessment_by_subject.get(subj.code,{}) or {}
        comp_sum=sum(float(comp.get(k,0) or 0) for k in ('multiple_assessment','subject_enrichment','portfolio'))
        internal=min(20.0,best2+comp_sum)
        final_val=None; final_max=80.0
        for e in finals:
            m=marks_by_subject.get((subj.code,e.id))
            if m and m.marks is not None:
                final_val=float(m.marks); final_max=float(m.max_marks or 80); break
        total=internal+final_val if final_val is not None else None
        grade=_grade((total/100)*100 if total is not None else 0) if total is not None else ''
        if total is not None:
            overall_total += total; overall_count += 1
        final_data.append([Paragraph(esc(subj.name),tiny),f'{best2:.1f}' if pt_scores else '—',f'{internal:.1f}' if (pt_scores or comp_sum) else '—',fmt_num(final_val) if final_val is not None else '—',f'{total:.1f}' if total is not None else '—',grade or '—'])
    if len(final_data)>1:
        story.append(_boxed_table(final_data,[46*mm,25*mm,25*mm,25*mm,27*mm,22*mm],1,6.1))
    if overall_count:
        overall_pct=overall_total/(overall_count*100)*100
        story.append(Spacer(1,5))
        story.append(Paragraph(f'<b>FINAL PERCENTAGE:</b> {overall_pct:.2f}% &nbsp;&nbsp; <b>Overall Total:</b> {overall_total:.1f} / {overall_count*100:.0f} &nbsp;&nbsp; <b>Grade:</b> {_grade(overall_pct)} &nbsp;&nbsp; <b>Result:</b> {"PASS" if overall_pct >= 33 else "REQUIRES ATTENTION"}',body))

    story.append(PageBreak())
    _header_story(story,student,session_name,logo_path)
    story.append(Paragraph(development_title,h))
    development_values=[]
    for label,key in [('Work Education','work_education'),('Art Education','art_education'),('Health & Physical Education','health_physical')]:
        if safe(co.get(key)): development_values.append((label,co.get(key)))
    for label,key in [('Discipline','discipline'),('Regularity','regularity'),('Punctuality','punctuality')]:
        if safe(dis.get(key)): development_values.append((label,dis.get(key)))
    if development_values:
        story.append(_boxed_table([[Paragraph('Area',small),Paragraph('Grade / Rating',small)]]+[[Paragraph(esc(a),small),Paragraph(esc(b),body)] for a,b in development_values],[120*mm,50*mm],1,7.0))
    else:
        story.append(Paragraph('No co-scholastic or discipline ratings have been entered.',section_note))

    health_rows=[]
    for term,keyh,keyw in [('Term I','term1_height','term1_weight'),('Term II','term2_height','term2_weight')]:
        if safe(health.get(keyh)) or safe(health.get(keyw)):
            health_rows.append([term,health.get(keyh,''),health.get(keyw,'')])
    if health_rows:
        story.append(Spacer(1,7)); story.append(Paragraph('Health & Physical Development',h))
        story.append(_boxed_table([[Paragraph(x,small) for x in ['Term','Height','Weight']]]+[[Paragraph(esc(a),small),Paragraph(esc(b),body),Paragraph(esc(c),body)] for a,b,c in health_rows],[60*mm,55*mm,55*mm],1,7.0))

    story.append(Spacer(1,7))
    story.append(Paragraph('Teacher & Principal Remarks',h))
    remark_rows=[]
    if safe(config.get('remarks')): remark_rows.append(['Teacher Remarks',config.get('remarks')])
    if safe(config.get('principal_remarks')): remark_rows.append(["Principal's Remarks",config.get('principal_remarks')])
    if remark_rows:
        story.append(_boxed_table([[Paragraph(esc(a),small),Paragraph(esc(b),body)] for a,b in remark_rows],[45*mm,125*mm],0,7.2))
    else:
        story.append(Paragraph('No remarks have been entered.',section_note))

    story.append(Spacer(1,7))
    progression=[]
    for label,key in [('Next Academic Session','next_academic_session'),('Session Begins','session_begins'),('Summer Break From','summer_break_from'),('School Re-opens','school_reopens')]:
        if safe(config.get(key)): progression.append([label,config.get(key)])
    if progression:
        story.append(Paragraph('Academic Progression & Calendar',h))
        story.append(_boxed_table([[Paragraph(esc(a),small),Paragraph(esc(b),body)] for a,b in progression],[60*mm,110*mm],0,7.0))

    story.append(Spacer(1,9))
    sig_img=_decode_signature(teacher_signature)
    sig_first=sig_img if sig_img else ''
    sig_data=[[sig_first,'',''],[teacher_signature_label,principal_signature_label,parent_signature_label]]
    sig=Table(sig_data,colWidths=[55*mm,55*mm,55*mm],rowHeights=[18*mm,8*mm])
    sig.setStyle(TableStyle([
        ('LINEBELOW',(0,0),(-1,0),0.45,colors.HexColor('#AAB4C1')),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('FONTSIZE',(0,1),(-1,1),7.0),
        ('TOPPADDING',(0,1),(-1,1),3)
    ]))
    story.append(sig)

    story.append(PageBreak())
    _header_story(story,student,session_name,logo_path)
    story.append(Paragraph(guide_title,h_center))
    story.append(Paragraph('This page summarizes the interpretation of the report and the academic grading scale used by the application.',section_note))
    story.append(_boxed_table([
        [Paragraph('Rule',small),Paragraph('School Policy / Explanation',small)],
        [Paragraph('Pass percentage',small),Paragraph(esc(pass_rule),body)],
        [Paragraph('Internal Assessment',small),Paragraph('Maximum 20 marks: best-two PT contribution plus Multiple Assessment, Subject Enrichment and Portfolio.',body)],
        [Paragraph('Final Examination',small),Paragraph('Shown using the examination maximum stored for the selected final examination.',body)],
        [Paragraph('Published results',small),Paragraph('Published results are protected from normal editing.',body)],
        [Paragraph('Principal authority',small),Paragraph('Final administrative decision rests with the Principal.',body)],
    ],[48*mm,122*mm],1,7.1))
    story.append(Spacer(1,9))
    story.append(Paragraph('Grading Scale',h))
    story.append(_boxed_table([
        [Paragraph('Marks Range',small),Paragraph('Grade',small),Paragraph('Interpretation',small)],
        ['91–100','A1','Outstanding'],['81–90','A2','Excellent'],['71–80','B1','Very Good'],['61–70','B2','Good'],
        ['51–60','C1','Satisfactory'],['41–50','C2','Developing'],['33–40','D','Needs Improvement'],['Below 33%','E','Requires Attention']
    ],[55*mm,30*mm,85*mm],1,7.0))
    story.append(Spacer(1,10))
    story.append(Paragraph('Record Notes',h))
    story.append(Paragraph('This report is generated from the student, attendance, examination, assessment, and report-card records available in the school management system at the time of generation.',body))

    doc.build(story)
    out.seek(0)
    return out

